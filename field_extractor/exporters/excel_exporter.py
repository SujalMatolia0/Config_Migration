"""
Excel Exporter for OSVC Field Extractor.

Generates structured Excel workbooks:
- Standard_Objects.xlsx
- Custom_Objects.xlsx
- Workspaces.xlsx & Workspaces_No_Ignored.xlsx
- Field_Catalog.xlsx & Field_Catalog_No_Ignored.xlsx
- Custom_Fields_Mapping.xlsx
"""

import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from mappings.field_id_mapping import get_mapped_rest_key
except ImportError:
    from field_extractor.mappings.field_id_mapping import get_mapped_rest_key

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

HEADER_FILL  = PatternFill("solid", fgColor="2E75B6")   # professional blue
ALT_ROW_FILL = PatternFill("solid", fgColor="DEEAF1")   # light blue-grey
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Calibri", size=10)
THIN_BORDER  = Border(
    left=Side(style="thin", color="9DC3E6"),
    right=Side(style="thin", color="9DC3E6"),
    top=Side(style="thin", color="9DC3E6"),
    bottom=Side(style="thin", color="9DC3E6"),
)

def _style_sheet(ws, headers):
    """Apply header style, freeze top row, add auto-filter."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"
    ws.row_dimensions[1].height = 34

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max(len(header) + 4, 14), 50)

def _write_rows(ws, rows, headers, start_row=2):
    """Write data rows with alternating row color."""
    for row_idx, row_data in enumerate(rows, start=start_row):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if fill:
                cell.fill = fill

def _safe_tab_name(name):
    """Sanitize string to valid Excel sheet name (max 31 chars)."""
    name = re.sub(r'[\\/*?:\[\]]', '_', name)
    return name[:31]

_TRIGGER_MAP = {
    "onnew":    "New",
    "onedit":   "Edit",
    "onsave":   "Save",
    "onalways": "Always",
    "ondelete": "Delete",
}

def _format_option(raw):
    """Converts OSVC workspace option strings to clean readable values."""
    if not raw or raw.strip() == "":
        return "No"

    raw = raw.strip()
    if raw.lower() in ("true", "yes", "1"):
        return "Yes"
    if raw.lower() in ("false", "no", "0"):
        return "No"

    segments = raw.split(";")
    triggers = []
    for seg in segments:
        seg = seg.strip()
        if not seg:
            continue
        if ":" in seg:
            trigger, condition = seg.split(":", 1)
            trigger   = trigger.strip().lower()
            condition = condition.strip().strip("~")
            label = _TRIGGER_MAP.get(trigger, trigger.capitalize())
            if condition == "any":
                triggers.append(label)
            else:
                triggers.append(f"{label} ({condition})")
        else:
            triggers.append(seg)

    if not triggers:
        return "No"
    if len(triggers) == 1:
        return f"Yes ({triggers[0]} Only)"
    return "Yes (" + " + ".join(triggers) + ")"

_SYSTEM_PACKAGES = {"oracleservicecloud", "rightnow", "system", "standard", "osvc", ""}

def _field_type_from_obj(field_dict):
    pkg  = (field_dict.get("package_name") or "").strip()
    name = (field_dict.get("field_name")   or "").strip()

    if pkg and pkg.lower() not in _SYSTEM_PACKAGES:
        return f"Custom ({pkg})"

    if "$" in name:
        pkg_part = name.split("$", 1)[0]
        return f"Custom ({pkg_part})"

    return "System Field" if field_dict.get("is_system_field", False) else "OSVC Standard"

def _field_type_from_ws_id(raw_field_id):
    if not raw_field_id:
        return "System Field"
    token = re.sub(r'^[^.]+\.', '', raw_field_id)
    token = re.sub(r'^CustomFields\.', '', token, flags=re.IGNORECASE)
    if "$" in token:
        pkg_part = token.split("$", 1)[0]
        return f"Custom ({pkg_part})"
    return "System Field"

def _obj_field_key(field_dict):
    pkg  = (field_dict.get("package_name") or "").strip()
    name = (field_dict.get("field_name")   or field_dict.get("field_id") or "").strip()
    if not pkg or pkg.lower() in _SYSTEM_PACKAGES or field_dict.get("is_system_field"):
        return name
    return f"{pkg}${name}" if name else pkg

def _ws_field_key(raw_field_id):
    if not raw_field_id:
        return ""
    token = str(raw_field_id).strip()
    # Strip workspace object prefixes (e.g. Incident., Contact.) but preserve REST keys like primarycontact.id
    token = re.sub(r'^(?:Incident|Contact|Organization|Task|Answer|Opportunity|Asset|ServiceCategory|Co)\.', '', token, flags=re.IGNORECASE)
    token = re.sub(r'^CustomFields\.', '', token, flags=re.IGNORECASE)
    return token.lower()

def _to_snake_case_col(field_name):
    if not field_name:
        return ""
    token = str(field_name).strip()
    token = re.sub(r'^(?:Incident|Contact|Organization|Task|Answer|Opportunity|Asset|ServiceCategory|Co)\.', '', token, flags=re.IGNORECASE)
    token = re.sub(r'^CustomFields\.', '', token, flags=re.IGNORECASE)
    if token.lower().startswith("c$"):
        prefix = "c$"
        rest = token[2:]
    else:
        prefix = ""
        rest = token
    rest_snake = re.sub(r'([a-z0-9])([A-Z])', r'\1_\2', rest).lower()
    return f"{prefix}{rest_snake}"

def _is_technical_code_label(label, field_name):
    if not label:
        return True
    lbl_clean = str(label).strip().lower().replace("_", "")
    fn_clean = str(field_name or "").strip().lower().replace("_", "")
    snake_clean = _to_snake_case_col(field_name).replace("_", "")
    bare_clean = snake_clean.replace("c$", "")
    return lbl_clean in (fn_clean, snake_clean, bare_clean, fn_clean.replace("c$", ""))

def _resolve_custom_field_label(xml_label, field_name, custom_fields_map, target_obj="Incident"):
    """
    Resolves Custom Field Label with fallback rules:
    1. First Priority: Human XML Label (if present and NOT a raw technical column key like 'C$ServiceAgency' or 'c$service_agency').
    2. Second Priority (Fallback): 'Name' column from Custom_Fields.xlsx (via custom_fields_map).
    """
    raw_col = (field_name or "").strip()
    clean_label = (xml_label or "").strip()

    # If XML label exists and is a REAL human label (not just a technical code like C$ServiceAgency), use XML label
    if clean_label and not _is_technical_code_label(clean_label, raw_col):
        return clean_label

    # Fallback to Custom_Fields.xlsx 'Name' column if available for custom fields
    if custom_fields_map and (raw_col or clean_label):
        snake = _to_snake_case_col(raw_col) or _to_snake_case_col(clean_label)
        col_low = raw_col.lower()
        col_c = snake if snake.startswith("c$") else f"c${snake}"
        col_bare = snake.replace("c$", "") if snake.startswith("c$") else snake
        col_no_us = col_c.replace("_", "")

        t_low = (target_obj or "incident").lower()
        t_sing = t_low[:-1] if t_low.endswith("s") else t_low

        candidates = [
            snake, col_c, col_bare, col_no_us, col_low,
            f"{t_sing}.${col_c}", f"{t_sing}.${col_bare}", f"{t_sing}.${snake}",
            f"{t_sing}.{col_c}", f"{t_sing}.{col_bare}", f"{t_sing}.{snake}",
            f"{t_low}.${col_c}", f"{t_low}.${col_bare}", f"{t_low}.${snake}",
            f"{t_low}.{col_c}", f"{t_low}.{col_bare}", f"{t_low}.{snake}"
        ]
        for c in candidates:
            if c in custom_fields_map and custom_fields_map[c].get("field_label"):
                return custom_fields_map[c]["field_label"].strip()

    return clean_label or raw_col

def _build_obj_index(objects_map):
    indexed = {}
    for obj_name, obj_data in (objects_map or {}).items():
        key_map = {}
        for f in obj_data.get("fields", []):
            field_key = _obj_field_key(f)
            key_map[field_key] = f
            name_lower = (f.get("field_name") or "").lower()
            if name_lower and name_lower not in key_map:
                key_map[name_lower] = f
            if "$" in name_lower:
                bare = name_lower.split("$", 1)[1]
                if bare and bare not in key_map:
                    key_map[bare] = f
        o_low = obj_name.lower()
        indexed[o_low] = key_map
        if o_low.endswith("s") and o_low[:-1] not in indexed:
            indexed[o_low[:-1]] = key_map
        elif not o_low.endswith("s") and (o_low + "s") not in indexed:
            indexed[o_low + "s"] = key_map
    return indexed

def _enrich_workspace_fields(ws_fields, objects_map, bound_object, standard_objects_map=None, custom_objects_map=None, custom_fields_map=None):
    if standard_objects_map is None and custom_objects_map is None:
        std_map = {}
        cst_map = {}
        for oname, odata in (objects_map or {}).items():
            if '.' in oname:
                cst_map[oname] = odata
            else:
                std_map[oname] = odata
                c_fields = [f for f in odata.get("fields", []) if f.get("package_name", "").lower() not in _SYSTEM_PACKAGES or "$" in f.get("field_name", "")]
                if c_fields:
                    cst_map[oname] = {"object_name": odata.get("object_name", oname), "fields": c_fields}
        std_indexed = _build_obj_index(std_map)
        cst_indexed = _build_obj_index(cst_map)
    else:
        std_indexed = _build_obj_index(standard_objects_map or {})
        cst_indexed = _build_obj_index(custom_objects_map or {})

    combined_indexed = _build_obj_index(objects_map or {})
    enriched = []

    for wf in ws_fields:
        target_obj = wf.get("target_object") or bound_object
        target_key = target_obj.lower()

        raw_id     = wf.get("raw_field_id", "") or wf.get("field_code", "")
        f_code     = wf.get("field_code", "")

        ws_key     = _ws_field_key(raw_id)
        ws_key_alt = _ws_field_key(f_code)

        is_custom_ws = ("$" in raw_id) or ("customfields" in raw_id.lower()) or ("$" in f_code)

        if is_custom_ws:
            primary_idx   = cst_indexed.get(target_key, {})
            secondary_idx = std_indexed.get(target_key, {})
        else:
            primary_idx   = std_indexed.get(target_key, {})
            secondary_idx = cst_indexed.get(target_key, {})

        comb_idx = combined_indexed.get(target_key, {})

        mapped_key     = _ws_field_key(get_mapped_rest_key(target_obj, raw_id))
        mapped_key_alt = _ws_field_key(get_mapped_rest_key(target_obj, f_code))

        matched = (
            primary_idx.get(mapped_key) or primary_idx.get(mapped_key_alt) or
            primary_idx.get(ws_key) or primary_idx.get(ws_key_alt) or
            secondary_idx.get(mapped_key) or secondary_idx.get(mapped_key_alt) or
            secondary_idx.get(ws_key) or secondary_idx.get(ws_key_alt) or
            comb_idx.get(mapped_key) or comb_idx.get(mapped_key_alt) or
            comb_idx.get(ws_key) or comb_idx.get(ws_key_alt)
        )

        if not matched and "." in ws_key:
            bare = ws_key.split(".")[-1]
            matched = primary_idx.get(bare) or secondary_idx.get(bare) or comb_idx.get(bare)

        cf_match = None
        if custom_fields_map and not matched:
            raw_clean  = (raw_id or f_code).lower()
            bare_clean = _ws_field_key(raw_clean)
            snake_clean = _to_snake_case_col(raw_id or f_code)
            t_low      = target_obj.lower()
            t_singular = t_low[:-1] if t_low.endswith("s") else t_low

            candidates = [
                snake_clean,
                raw_clean,
                raw_clean.replace("_", ""),
                bare_clean,
                bare_clean.replace("_", ""),
                f"{t_singular}.${snake_clean}",
                f"{t_singular}.${bare_clean}",
                f"{t_singular}.${bare_clean.replace('_', '')}",
                f"{t_singular}.{snake_clean}",
                f"{t_singular}.{bare_clean}",
                f"{t_singular}.{bare_clean.replace('_', '')}",
                f"{t_low}.${snake_clean}",
                f"{t_low}.${bare_clean}",
                f"{t_low}.${bare_clean.replace('_', '')}",
                f"{t_low}.{snake_clean}",
                f"{t_low}.{bare_clean}",
                f"{t_low}.{bare_clean.replace('_', '')}",
            ]
            for c in candidates:
                if c in custom_fields_map:
                    cf_match = custom_fields_map[c]
                    break

        if matched:
            data_type     = matched.get("data_type", "Text")
            ftype         = _field_type_from_obj(matched)
            is_sys        = "Yes" if matched.get("is_system_field") else "No"
            raw_pkg       = (matched.get("package_name") or "").strip()
            if not raw_pkg or raw_pkg.lower() in _SYSTEM_PACKAGES or matched.get("is_system_field"):
                pkg = ""
            else:
                pkg = raw_pkg
            is_nullable   = "Yes" if matched.get("is_nullable") else "No"
            is_lookup     = "Yes" if matched.get("is_lookup")   else "No"
            is_readonly_s = "Yes" if matched.get("is_readonly") else "No"
            is_list       = "Yes" if matched.get("is_list")     else "No"
            is_autoupdate = "Yes" if matched.get("is_autoupdate") else "No"
            is_sequence   = "Yes" if matched.get("is_sequence")   else "No"
            max_len       = str(matched.get("max_length", "-"))
            desc          = matched.get("description", "")
            avail_get  = ("-" if "is_available_get"  not in matched else "Yes" if matched["is_available_get"]  else "No")
            avail_post = ("-" if "is_available_post" not in matched else "Yes" if matched["is_available_post"] else "No")
            avail_patch = ("-" if "is_available_patch" not in matched else "Yes" if matched["is_available_patch"] else "No")
            is_deprec  = ("-" if "is_deprecated" not in matched else "Yes" if matched["is_deprecated"] else "No")
            is_enum_str = "-" if "isEnumerable" not in matched else "Yes" if matched.get("isEnumerable") else "No"
            minimum   = "-" if "minimum" not in matched else str(matched["minimum"])
            maximum   = "-" if "maximum" not in matched else str(matched["maximum"])
            ref_val   = str(matched.get("$ref") or matched.get("ref") or matched.get("ref_url") or "-")
            items_val = matched.get("items")
            if "items" not in matched and not matched.get("is_list"):
                items_str = "-"
            elif isinstance(items_val, (dict, list)):
                items_str = str(items_val)
            elif matched.get("is_list"):
                items_str = "IsList: Yes"
            else:
                items_str = str(items_val) if items_val is not None else "-"
            pattern     = str(matched.get("pattern", "-"))

            fname = matched.get("field_name", "")
            raw_low = (raw_id or f_code).lower()
            # Preserve subfield paths from XML (e.g. name.first, name.last) when matching parent objects like name
            if ("." in raw_low) and not fname.lower().startswith(raw_low):
                fn_low = fname.lower()
                if raw_low.startswith(f"{fn_low}."):
                    fname = raw_id or f_code
                    if data_type in ("Object", "PersonName", "NamedID", "Compound"):
                        data_type = "Text"

            obj_field_key = fname if not pkg else f"{pkg}${fname}"
            unmapped_category = ""
        elif cf_match:
            data_type     = cf_match.get("data_type", "Custom Field")
            ftype         = f"Custom Field ({cf_match.get('column_name')})"
            is_sys        = "No"
            pkg           = ""
            is_nullable   = "Yes"
            is_lookup     = "No"
            is_readonly_s = "No"
            is_list       = "No"
            is_autoupdate = "No"
            is_sequence   = "No"
            max_len       = str(cf_match.get("field_size") or "-")
            desc          = f"Custom Field ID: {cf_match.get('custom_field_id')} | Flags: {cf_match.get('flags', '')}"
            avail_get     = "-"
            avail_post    = "-"
            avail_patch   = "-"
            is_deprec     = "-"
            is_enum_str   = "Yes" if cf_match.get("menu_items") else "No"
            minimum       = "-"
            maximum       = "-"
            ref_val       = "-"
            items_str     = cf_match.get("menu_items_str") or "-"
            pattern       = "-"
            obj_field_key = cf_match.get("column_name") or raw_id
            unmapped_category = ""
            if not wf.get("field_label") and cf_match.get("field_label"):
                wf["field_label"] = cf_match.get("field_label")
        else:
            data_type     = "-"
            ftype         = _field_type_from_ws_id(raw_id or f_code)
            is_sys        = "-"
            pkg           = ftype.split("(")[1].rstrip(")") if "Custom (" in ftype else ""
            is_nullable   = "-"
            is_lookup     = "-"
            is_readonly_s = "-"
            is_list       = "-"
            is_autoupdate = "-"
            is_sequence   = "-"
            max_len       = "-"
            desc          = ""
            avail_get     = "-"
            avail_post    = "-"
            avail_patch   = "-"
            is_deprec     = "-"
            is_enum_str   = "-"
            minimum       = "-"
            maximum       = "-"
            ref_val       = "-"
            items_str     = "-"
            pattern       = "-"
            obj_field_key = raw_id

            fid_low = (raw_id or f_code).lower()
            if fid_low.startswith("rulectx.") or fid_low == "rulestate":
                unmapped_category = "Business Rule Context Variable"
            elif fid_low.endswith(".invalid"):
                unmapped_category = "UI Validation Control"
            elif fid_low in ("dormant", "linesperpage", "searchtext", "searchtype", "state", "passwordencrypt"):
                unmapped_category = "Workspace UI Control"
            elif fid_low in ("mamailtype", "maorgname", "totrev"):
                unmapped_category = "Legacy Workspace Field"
            else:
                unmapped_category = "Unmapped Workspace Control"

        xml_lbl = wf.get("field_label") or (matched.get("field_label") if matched else "")
        f_name_raw = (matched.get("field_name") if matched else (cf_match.get("column_name") if cf_match else raw_id or f_code))
        flabel = _resolve_custom_field_label(xml_lbl, f_name_raw, custom_fields_map, target_obj)

        item = dict(wf)
        item.update({
            "field_label":        flabel,
            "is_unmapped":        matched is None and cf_match is None,
            "unmapped_category":  unmapped_category if (matched is None and cf_match is None) else "",
            "target_object":      target_obj,
            "obj_field_key":      obj_field_key,
            "data_type":          data_type,
            "field_type":         ftype,
            "is_system_field":    is_sys,
            "package_name":       pkg,
            "is_nullable":        is_nullable,
            "is_lookup":          is_lookup,
            "is_readonly_schema": is_readonly_s,
            "is_list":            is_list,
            "is_autoupdate":      is_autoupdate,
            "is_sequence":        is_sequence,
            "max_length":         max_len,
            "description":        desc,
            "avail_get":          avail_get,
            "avail_post":         avail_post,
            "avail_patch":        avail_patch,
            "is_deprecated":      is_deprec,
            "is_enumerable":      is_enum_str,
            "minimum":            minimum,
            "maximum":            maximum,
            "ref":                ref_val,
            "items":              items_str,
            "pattern":            pattern,
            "required_fmt":       _format_option(wf.get("required_option", "")),
            "readonly_fmt":       _format_option(wf.get("readonly_option", "")),
        })
        enriched.append(item)

    return enriched

def _write_ignored_fields_tab(wb, unmapped_records):
    if not unmapped_records:
        return

    headers = [
        "Workspace Name", "Bound Object", "Target Object",
        "Field ID", "Field Label", "Location / Tab", "Reason / Category"
    ]
    sheet = wb.create_sheet(title="Ignored_Fields")
    _style_sheet(sheet, headers)

    rows = []
    for r in unmapped_records:
        rows.append([
            r["workspace_name"],
            r["bound_object"],
            r["target_object"],
            r.get("raw_field_id") or r.get("field_code") or "",
            r.get("field_label", ""),
            r.get("location_tab", ""),
            r.get("unmapped_category", "Unmapped Workspace Control")
        ])
    _write_rows(sheet, rows, headers)

def write_workspaces_excel(parsed_workspaces, objects_map, output_path, include_ignored_tab=True, custom_fields_map=None, simplify_attributes=False):
    if not parsed_workspaces:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "No_Workspaces"
        ws.cell(row=1, column=1, value="No workspace definitions extracted.")
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        return output_path

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if simplify_attributes:
        headers = [
            "Bound Object", "Target Object", "Field Name",
            "Field Label", "Location / Tab", "Required", "Read Only",
            "Data Type", "Field Type", "Is Nullable", "Is Lookup", "Is List", "Is Auto Update", "Max Length",
            "Description"
        ]
    else:
        headers = [
            "Bound Object", "Target Object", "Field Name",
            "Field Label", "Location / Tab", "Required", "Read Only",
            "Data Type", "Field Type", "Is Nullable", "Is Lookup", "Is List", "Is Auto Update", "Max Length",
            "Is Available GET", "Is Available POST", "Is Available PATCH", "Is Deprecated", "Description"
        ]

    all_unmapped = []
    for ws_data in parsed_workspaces:
        ws_name      = ws_data["workspace_name"]
        bound_object = ws_data.get("bound_object", "Contact")
        enriched     = _enrich_workspace_fields(ws_data.get("fields", []), objects_map, bound_object, custom_fields_map=custom_fields_map)

        sheet = wb.create_sheet(title=_safe_tab_name(ws_name))
        _style_sheet(sheet, headers)

        rows = []
        for item in enriched:
            if item.get("is_unmapped"):
                rec = dict(item)
                rec["workspace_name"] = ws_name
                all_unmapped.append(rec)

            if simplify_attributes:
                rows.append([
                    item["bound_object"],
                    item["target_object"],
                    item["obj_field_key"],
                    item["field_label"],
                    item["location_tab"],
                    item["required_fmt"],
                    item["readonly_fmt"],
                    item["data_type"],
                    item["field_type"],
                    item["is_nullable"],
                    item["is_lookup"],
                    item["is_list"],
                    item["is_autoupdate"],
                    item["max_length"],
                    item.get("description", ""),
                ])
            else:
                rows.append([
                    item["bound_object"],
                    item["target_object"],
                    item["obj_field_key"],
                    item["field_label"],
                    item["location_tab"],
                    item["required_fmt"],
                    item["readonly_fmt"],
                    item["data_type"],
                    item["field_type"],
                    item["is_nullable"],
                    item["is_lookup"],
                    item["is_list"],
                    item["is_autoupdate"],
                    item["max_length"],
                    item.get("avail_get", "-"),
                    item.get("avail_post", "-"),
                    item.get("avail_patch", "-"),
                    item.get("is_deprecated", "-"),
                    item.get("description", ""),
                ])

        _write_rows(sheet, rows, headers)

    if include_ignored_tab:
        _write_ignored_fields_tab(wb, all_unmapped)
    wb.save(output_path)
    return output_path

def write_objects_excel(objects_map, output_path, simplify_attributes=False, custom_fields_map=None):
    if not objects_map:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "No_Objects"
        ws.cell(row=1, column=1, value="No object schemas extracted.")
        wb.save(output_path)
        return output_path

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if simplify_attributes:
        headers = [
            "Field Name", "Field Label", "Data Type", "Field Type",
            "Is System Field", "Package Name", "Is Nullable", "Is Lookup",
            "Is Read Only", "Max Length", "Description",
            "Is Enumerable", "$Ref", "Items", "Pattern"
        ]
    else:
        headers = [
            "Field Name", "Field Label", "Data Type", "Field Type",
            "Is System Field", "Package Name", "Is Nullable", "Is Lookup",
            "Is Read Only", "Max Length", "Description",
            "Is Available GET", "Is Available POST", "Is Available PATCH",
            "Is Deprecated", "Is Enumerable", "Minimum", "Maximum",
            "$Ref", "Items", "Pattern"
        ]

    for obj_name, obj_data in objects_map.items():
        display_name = obj_data.get("object_name", obj_name)
        sheet = wb.create_sheet(title=_safe_tab_name(display_name))
        _style_sheet(sheet, headers)

        rows = []
        for of in obj_data.get("fields", []):
            key = _obj_field_key(of)
            raw_pkg = (of.get("package_name") or "").strip()
            display_pkg = "" if (not raw_pkg or raw_pkg.lower() in _SYSTEM_PACKAGES or of.get("is_system_field")) else raw_pkg
            xml_lbl = of.get("field_label", "")
            f_name_raw = of.get("field_name", "")
            flabel = _resolve_custom_field_label(xml_lbl, f_name_raw, custom_fields_map, target_obj=display_name)
            is_enum_val = of.get("isEnumerable")
            is_enum_str = "Yes" if is_enum_val is True else ("No" if is_enum_val is False else (str(is_enum_val) if is_enum_val is not None else "-"))
            items_val = of.get("items")
            items_str = str(items_val) if isinstance(items_val, (dict, list)) else ("IsList: Yes" if of.get("is_list") else (str(items_val) if items_val is not None else "-"))
            ref_val = of.get("$ref") or of.get("ref") or of.get("ref_url") or "-"

            if simplify_attributes:
                row = [
                    key,
                    flabel,
                    of.get("data_type", ""),
                    _field_type_from_obj(of),
                    "Yes" if of.get("is_system_field") else "No",
                    display_pkg,
                    "Yes" if of.get("is_nullable") else "No",
                    "Yes" if of.get("is_lookup")   else "No",
                    "Yes" if of.get("is_readonly") else "No",
                    str(of.get("max_length", "-")),
                    of.get("description", ""),
                    is_enum_str,
                    str(ref_val),
                    items_str,
                    str(of.get("pattern", "-"))
                ]
            else:
                row = [
                    key,
                    flabel,
                    of.get("data_type", ""),
                    _field_type_from_obj(of),
                    "Yes" if of.get("is_system_field") else "No",
                    display_pkg,
                    "Yes" if of.get("is_nullable") else "No",
                    "Yes" if of.get("is_lookup")   else "No",
                    "Yes" if of.get("is_readonly") else "No",
                    str(of.get("max_length", "-")),
                    of.get("description", ""),
                    ("-" if "is_available_get"   not in of else "Yes" if of["is_available_get"]   else "No"),
                    ("-" if "is_available_post"  not in of else "Yes" if of["is_available_post"]  else "No"),
                    ("-" if "is_available_patch" not in of else "Yes" if of["is_available_patch"] else "No"),
                    ("-" if "is_deprecated"      not in of else "Yes" if of["is_deprecated"]      else "No"),
                    is_enum_str,
                    str(of.get("minimum", "-")),
                    str(of.get("maximum", "-")),
                    str(ref_val),
                    items_str,
                    str(of.get("pattern", "-"))
                ]
            rows.append(row)

        _write_rows(sheet, rows, headers)

    wb.save(output_path)
    return output_path

def write_combined_excel(parsed_workspaces, objects_map, output_path, include_ignored_tab=True, custom_fields_map=None, simplify_attributes=True):
    if not parsed_workspaces:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "No_Workspaces"
        ws.cell(row=1, column=1, value="No workspace definitions extracted.")
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        return output_path

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if simplify_attributes:
        headers = [
            "Bound Object", "Target Object", "Field Name", "Field Label",
            "Workspace Tab", "Required (In Layout)", "Read Only (In Layout)",
            "Data Type", "Field Type", "Is System Field", "Package Name",
            "Is Nullable", "Is Lookup", "Is Read Only (Schema)", "Max Length",
            "Description", "Is Enumerable", "$Ref", "Items",
            "Pattern", "Is List", "Is Auto Update", "In Workspace Layout"
        ]
    else:
        headers = [
            "Bound Object", "Target Object", "Field Name", "Field Label",
            "Workspace Tab", "Required (In Layout)", "Read Only (In Layout)",
            "Data Type", "Field Type", "Is System Field", "Package Name",
            "Is Nullable", "Is Lookup", "Is Read Only (Schema)", "Max Length",
            "Description", "Is Available GET", "Is Available POST", "Is Available PATCH",
            "Is Deprecated", "Is Enumerable", "Minimum", "Maximum", "$Ref", "Items",
            "Pattern", "Is List", "Is Auto Update", "In Workspace Layout"
        ]

    all_unmapped = []
    for ws_data in parsed_workspaces:
        ws_name      = ws_data["workspace_name"]
        bound_object = ws_data.get("bound_object", "Contact")
        enriched     = _enrich_workspace_fields(ws_data.get("fields", []), objects_map, bound_object, custom_fields_map=custom_fields_map)

        sheet = wb.create_sheet(title=_safe_tab_name(ws_name))
        _style_sheet(sheet, headers)

        rows = []
        for item in enriched:
            if item.get("is_unmapped"):
                rec = dict(item)
                rec["workspace_name"] = ws_name
                all_unmapped.append(rec)

            if simplify_attributes:
                rows.append([
                    item["bound_object"],
                    item["target_object"],
                    item["obj_field_key"],
                    item["field_label"],
                    item["location_tab"],
                    item["required_fmt"],
                    item["readonly_fmt"],
                    item["data_type"],
                    item["field_type"],
                    item["is_system_field"],
                    item["package_name"],
                    item["is_nullable"],
                    item["is_lookup"],
                    item["is_readonly_schema"],
                    item["max_length"],
                    item["description"],
                    item["is_enumerable"],
                    item["ref"],
                    item["items"],
                    item["pattern"],
                    item["is_list"],
                    item["is_autoupdate"],
                    "Yes",
                ])
            else:
                rows.append([
                    item["bound_object"],
                    item["target_object"],
                    item["obj_field_key"],
                    item["field_label"],
                    item["location_tab"],
                    item["required_fmt"],
                    item["readonly_fmt"],
                    item["data_type"],
                    item["field_type"],
                    item["is_system_field"],
                    item["package_name"],
                    item["is_nullable"],
                    item["is_lookup"],
                    item["is_readonly_schema"],
                    item["max_length"],
                    item["description"],
                    item["avail_get"],
                    item["avail_post"],
                    item["avail_patch"],
                    item["is_deprecated"],
                    item["is_enumerable"],
                    item["minimum"],
                    item["maximum"],
                    item["ref"],
                    item["items"],
                    item["pattern"],
                    item["is_list"],
                    item["is_autoupdate"],
                    "Yes",
                ])

        _write_rows(sheet, rows, headers)

    if include_ignored_tab:
        _write_ignored_fields_tab(wb, all_unmapped)
    wb.save(output_path)
    return output_path

def write_field_catalog_excel(parsed_workspaces, objects_map, output_path, include_ignored_tab=True, custom_fields_map=None, simplify_attributes=True):
    return write_combined_excel(parsed_workspaces, objects_map, output_path, include_ignored_tab=include_ignored_tab, custom_fields_map=custom_fields_map, simplify_attributes=simplify_attributes)

def write_custom_fields_mapping_excel(cf_map, output_path):
    """
    Generates a dedicated Custom Fields & Menu Options Excel mapping workbook:
    - Sheet 1: Custom_Fields_Catalog (Table, Column Name, Field Label, Custom Field ID, Data Type, Size, Menu Items Count, Menu Items, Flags)
    - Sheet 2: Menu_Custom_Fields (Menu Fields & Dropdown Options)
    - Sheets 3+: Table-specific sheets (Incident_Custom_Fields, Contact_Custom_Fields, etc.)
    """
    if not cf_map:
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    headers = [
        "Table Name", "Column Name", "Field Label", "Custom Field ID",
        "Data Type", "Field Size", "Menu Items Count", "Menu Items / Options", "Flags & Permissions"
    ]

    seen = set()
    unique_fields = []
    for k, v in cf_map.items():
        unique_key = f"{v.get('table_name')}::{v.get('column_name')}"
        if unique_key not in seen:
            seen.add(unique_key)
            unique_fields.append(v)

    unique_fields.sort(key=lambda x: (x.get("table_name", ""), x.get("field_label", "")))

    # Master Sheet: All Custom Fields
    master_sheet = wb.create_sheet(title="Custom_Fields_Catalog")
    _style_sheet(master_sheet, headers)

    rows = []
    for f in unique_fields:
        menu_items = f.get("menu_items", [])
        menu_cnt = len(menu_items)
        menu_str = f.get("menu_items_str") or (":".join(menu_items) if menu_items else "N/A")

        rows.append([
            f.get("table_name", "Incident"),
            f.get("column_name", ""),
            f.get("field_label", ""),
            f.get("custom_field_id") or "N/A",
            f.get("data_type", "Text"),
            f.get("field_size") or "—",
            menu_cnt if menu_cnt > 0 else "—",
            menu_str,
            f.get("flags") or "Standard Access"
        ])

    _write_rows(master_sheet, rows, headers)

    # Sheet 2: Menu Custom Fields Only
    menu_fields = [f for f in unique_fields if f.get("data_type") == "Menu" or f.get("menu_items")]
    if menu_fields:
        menu_sheet = wb.create_sheet(title="Menu_Custom_Fields")
        _style_sheet(menu_sheet, headers)
        menu_rows = []
        for f in menu_fields:
            menu_items = f.get("menu_items", [])
            menu_cnt = len(menu_items)
            menu_str = f.get("menu_items_str") or ":".join(menu_items)

            menu_rows.append([
                f.get("table_name", "Incident"),
                f.get("column_name", ""),
                f.get("field_label", ""),
                f.get("custom_field_id") or "N/A",
                "Menu",
                "—",
                menu_cnt,
                menu_str,
                f.get("flags") or "Standard Access"
            ])
        _write_rows(menu_sheet, menu_rows, headers)

    # Sheet per Table: Incident, Contact, Answer
    tables_group = {}
    for f in unique_fields:
        tbl = f.get("table_name", "Incident")
        tables_group.setdefault(tbl, []).append(f)

    for tbl, t_fields in sorted(tables_group.items()):
        sheet_title = _safe_tab_name(f"{tbl}_Custom_Fields")
        t_sheet = wb.create_sheet(title=sheet_title)
        _style_sheet(t_sheet, headers)
        t_rows = []
        for f in t_fields:
            menu_items = f.get("menu_items", [])
            menu_cnt = len(menu_items)
            menu_str = f.get("menu_items_str") or (":".join(menu_items) if menu_items else "N/A")

            t_rows.append([
                f.get("table_name", tbl),
                f.get("column_name", ""),
                f.get("field_label", ""),
                f.get("custom_field_id") or "N/A",
                f.get("data_type", "Text"),
                f.get("field_size") or "—",
                menu_cnt if menu_cnt > 0 else "—",
                menu_str,
                f.get("flags") or "Standard Access"
            ])
        _write_rows(t_sheet, t_rows, headers)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path

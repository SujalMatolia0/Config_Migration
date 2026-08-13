import os
import sys
import tempfile
import zipfile
import io
import traceback
import openpyxl
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename

# Add field_extractor directory to Python path if needed
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
if CURRENT_DIR not in sys.path:
    sys.path.insert(0, CURRENT_DIR)

from parsers.workspace_parser import parse_workspace_xml
from parsers.object_parser import parse_object_xml
from parsers.custom_field_excel_parser import parse_custom_fields_excel
from exporters.excel_exporter import (
    _enrich_workspace_fields,
    _field_type_from_obj,
    _field_type_from_ws_id,
    _obj_field_key,
    _format_option,
    write_workspaces_excel,
    write_objects_excel,
    write_combined_excel,
    write_custom_fields_mapping_excel,
)
from fetchers.osvc_rest_fetcher import fetch_standard_objects_via_rest

# Load connection config if available
try:
    from config import BASE_URL, USERNAME, PASSWORD as CONFIG_PASSWORD
    _cfg_host = BASE_URL
    _cfg_user = USERNAME
    _cfg_pass = CONFIG_PASSWORD
except ImportError:
    _cfg_host = ""
    _cfg_user = ""
    _cfg_pass = ""

app = Flask(__name__, template_folder=os.path.join(CURRENT_DIR, "templates"))
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB max upload limit

def load_standard_objects_from_excel():
    """Loads extracted standard objects directly from Standard_Objects.xlsx if present."""
    candidate_paths = [
        os.path.join(CURRENT_DIR, "schemas", "Standard_Objects.xlsx"),
        os.path.join(CURRENT_DIR, "input", "Standard_Objects.xlsx"),
        os.path.join(CURRENT_DIR, "sample_inputs", "Standard_Objects.xlsx"),
        os.path.join(CURRENT_DIR, "results", "Standard_Objects.xlsx"),
        os.path.join(CURRENT_DIR, "results", "standard_objects.xlsx"),
        os.path.join(CURRENT_DIR, "..", "results", "field_extractor", "Standard_Objects.xlsx"),
        os.path.join(CURRENT_DIR, "..", "input", "Standard_Objects.xlsx"),
    ]
    xlsx_path = next((p for p in candidate_paths if os.path.exists(p)), None)
    if not xlsx_path:
        add_log("Standard_Objects.xlsx not found in input or results folders.", "WARNING")
        return {}
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        std_map = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h or '') for h in rows[0]]
            fields = []
            for r in rows[1:]:
                if not any(r):
                    continue
                r_dict = dict(zip(headers, r))
                # Full Field Key from REST (e.g. 'primarycontact.id', 'subject', 'statuswithtype')
                fk = str(r_dict.get('Field Key', '') or '').strip()
                # Bare name: after '$' for custom fields, else the key itself
                fn_bare = fk.split('$')[-1] if '$' in fk else fk
                # Package: custom if key contains '$', else system
                pkg = r_dict.get('Package Name', '') or ''
                if '$' in fk and not pkg:
                    pkg = fk.split('$')[0].upper()
                entry = {
                    'field_id':   fk,       # full REST key used for lookup matching
                    'field_name': fk,       # must be full key so _build_obj_index indexes correctly
                    'field_name_bare': fn_bare,
                    'field_label': r_dict.get('Field Label', ''),
                    'data_type': r_dict.get('Data Type', ''),
                    'is_system_field': r_dict.get('Is System Field') == 'Yes',
                    'package_name': pkg,
                    'is_nullable': r_dict.get('Is Nullable') == 'Yes',
                    'is_lookup': r_dict.get('Is Lookup') == 'Yes',
                    'is_readonly': r_dict.get('Is Read Only') == 'Yes',
                    'max_length': r_dict.get('Max Length', '-'),
                    'description': r_dict.get('Description', ''),
                    'is_available_get':   r_dict.get('Is Available GET') == 'Yes',
                    'is_available_post':  r_dict.get('Is Available POST') == 'Yes',
                    'is_available_patch': r_dict.get('Is Available PATCH') == 'Yes',
                    'is_deprecated': r_dict.get('Is Deprecated') == 'Yes',
                    'isEnumerable': r_dict.get('Is Enumerable'),
                    'minimum': r_dict.get('Minimum'),
                    'maximum': r_dict.get('Maximum'),
                    '$ref': r_dict.get('$Ref'),
                    'items': r_dict.get('Items'),
                    'pattern': r_dict.get('Pattern'),
                }
                fields.append(entry)
                # Also add an alias entry with the bare name so lookups like 'subject' still work
                if fn_bare and fn_bare != fk:
                    alias = dict(entry)
                    alias['field_name'] = fn_bare
                    alias['field_id']   = fn_bare
                    fields.append(alias)
            std_map[sheet_name.lower()] = {
                'object_name': sheet_name,
                'fields': fields
            }
        return std_map
    except Exception as e:
        add_log(f"Error reading standard_objects.xlsx: {e}", "WARNING")
        return {}

def _save_std_cache(fetched_objects):
    """Saves fetched standard objects schema map to Standard_Objects.xlsx in results directory."""
    if not fetched_objects:
        return
    try:
        out_dir = os.path.join(CURRENT_DIR, "results")
        os.makedirs(out_dir, exist_ok=True)
        xlsx_path = os.path.join(out_dir, "Standard_Objects.xlsx")
        write_objects_excel(fetched_objects, xlsx_path)
    except Exception as e:
        add_log(f"Error saving standard objects cache: {e}", "WARNING")

# Global store for current session outputs
RESULTS_CACHE = {
    "workspaces": [],
    "standard_objects_map": {},
    "custom_objects_map": {},
    "combined_objects_map": {},
    "output_dir": None,
    "summary": {},
    "logs": []
}

def merge_objects_maps(standard_map, custom_map):
    """
    Merges standard objects map (from REST API) and custom objects map (from XMLs).
    Allows dual-lookup enrichment for workspace fields.
    """
    merged = {}
    for obj_key, obj_data in (standard_map or {}).items():
        merged[obj_key] = {
            "object_name": obj_data.get("object_name", obj_key),
            "fields": list(obj_data.get("fields", []))
        }
    for obj_key, obj_data in (custom_map or {}).items():
        if obj_key in merged:
            existing_fields = merged[obj_key]["fields"]
            existing_keys = {f.get("field_name", "").lower() for f in existing_fields}
            for cf in obj_data.get("fields", []):
                if cf.get("field_name", "").lower() not in existing_keys:
                    existing_fields.append(cf)
        else:
            merged[obj_key] = {
                "object_name": obj_data.get("object_name", obj_key),
                "fields": list(obj_data.get("fields", []))
            }
    return merged


def _build_workspaces_preview(workspaces, objects_map, custom_fields_map=None):
    preview_workspaces = []
    for ws_data in (workspaces or []):
        bound_obj = ws_data.get("bound_object", "Contact")
        enriched = _enrich_workspace_fields(
            ws_data.get("fields", []),
            objects_map,
            bound_obj,
            custom_fields_map=custom_fields_map
        )
        rows = []
        for item in enriched:
            rows.append({
                "bound_object":      item.get("bound_object") or "-",
                "target_object":     item.get("target_object") or "-",
                "object_field_name": item.get("obj_field_key") or "-",
                "field_label":       item.get("field_label") or "-",
                "location_tab":      item.get("location_tab") or "-",
                "required":          item.get("required_fmt") or "-",
                "readonly":          item.get("readonly_fmt") or "-",
                "data_type":         item.get("data_type") or "-",
                "field_type":        item.get("field_type") or "-",
                "is_nullable":       item.get("is_nullable") or "-",
                "is_lookup":         item.get("is_lookup") or "-",
                "max_length":        str(item.get("max_length") if item.get("max_length") is not None else "-"),
            })
        preview_workspaces.append({
            "workspace_name": ws_data.get("workspace_name", "Workspace"),
            "bound_object":   bound_obj,
            "field_count":    len(rows),
            "rows":           rows
        })
    return preview_workspaces


def _build_combined_preview(workspaces, objects_map, custom_fields_map=None):
    preview_combined = []
    for ws_data in (workspaces or []):
        bound_obj = ws_data.get("bound_object", "Contact")
        enriched = _enrich_workspace_fields(
            ws_data.get("fields", []),
            objects_map,
            bound_obj,
            custom_fields_map=custom_fields_map
        )
        rows = []
        for item in enriched:
            rows.append({
                "bound_object":       item.get("bound_object") or "-",
                "target_object":      item.get("target_object") or "-",
                "object_field_name":  item.get("obj_field_key") or "-",
                "field_label":        item.get("field_label") or "-",
                "workspace_tab":      item.get("location_tab") or "-",
                "location_tab":       item.get("location_tab") or "-",
                "required":           item.get("required_fmt") or "-",
                "readonly":           item.get("readonly_fmt") or "-",
                "data_type":          item.get("data_type") or "-",
                "field_type":         item.get("field_type") or "-",
                "is_system_field":    item.get("is_system_field") or "-",
                "package_name":       item.get("package_name") or "-",
                "is_nullable":        item.get("is_nullable") or "-",
                "is_lookup":          item.get("is_lookup") or "-",
                "is_readonly_schema": item.get("is_readonly_schema") or "-",
                "max_length":         str(item.get("max_length") if item.get("max_length") is not None else "-"),
                "description":        item.get("description") or "-",
                "avail_get":          item.get("avail_get") or "-",
                "avail_post":         item.get("avail_post") or "-",
                "avail_patch":        item.get("avail_patch") or "-",
                "is_deprecated":      item.get("is_deprecated") or "-",
                "is_enumerable":      item.get("is_enumerable") or "-",
                "minimum":            str(item.get("minimum") if item.get("minimum") is not None else "-"),
                "maximum":            str(item.get("maximum") if item.get("maximum") is not None else "-"),
                "ref":                str(item.get("ref") if item.get("ref") is not None else "-"),
                "items":              str(item.get("items") if item.get("items") is not None else "-"),
                "pattern":            str(item.get("pattern") if item.get("pattern") is not None else "-"),
                "is_list":            item.get("is_list") or "-",
                "is_autoupdate":      item.get("is_autoupdate") or "-",
                "in_layout":          "Yes"
            })
        preview_combined.append({
            "workspace_name": ws_data.get("workspace_name", "Workspace"),
            "bound_object":   bound_obj,
            "field_count":    len(rows),
            "rows":           rows
        })
    return preview_combined

def add_log(msg, level="INFO"):
    t_str = datetime.now().strftime("%H:%M:%S")
    entry = f"[{t_str}] [{level}] {msg}"
    print(entry)
    RESULTS_CACHE["logs"].append(entry)

@app.route("/api/config", methods=["GET"])
def get_config():
    """Returns saved connection config for auto-filling the UI form. Password is masked."""
    return jsonify({
        "host": _cfg_host,
        "username": _cfg_user,
        "password": _cfg_pass  # sent only to pre-fill local browser form
    })

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/logs", methods=["GET"])
def get_logs():
    return jsonify({"logs": RESULTS_CACHE["logs"]})

@app.route("/api/logs/clear", methods=["POST"])
def clear_logs():
    RESULTS_CACHE["logs"] = []
    add_log("Log history cleared.", "INFO")
    return jsonify({"success": True})

def _scan_input_directories(dirs):
    ws_files = []
    obj_files = []
    for d in dirs:
        if not d or not os.path.exists(d):
            continue
        for root_dir, _, files in os.walk(d):
            for fname in sorted(files):
                if fname.endswith(".xml") and not fname.startswith("."):
                    full_path = os.path.join(root_dir, fname)
                    try:
                        with open(full_path, "r", encoding="utf-8", errors="ignore") as f_handle:
                            content = f_handle.read(2048)
                        if "<CustomObject" in content or "<Fields>" in content:
                            obj_files.append(full_path)
                        elif "<Workspace" in content:
                            ws_files.append(full_path)
                    except Exception:
                        pass
    return ws_files, obj_files

@app.route("/api/load_sample", methods=["POST"])
def load_sample():
    """Loads all input files directly from field_extractor/input and sample_inputs."""
    add_log("Loading input files from field_extractor/input...", "INFO")
    fe_input_dir = os.path.join(CURRENT_DIR, "input")
    sample_dir = os.path.join(CURRENT_DIR, "sample_inputs")
    main_input_dir = os.path.abspath(os.path.join(CURRENT_DIR, "..", "input"))

    ws_files, obj_files = _scan_input_directories([fe_input_dir, sample_dir, main_input_dir])

    # Search for custom_fields Excel files across field_extractor/input and fallback paths
    f1_paths = [
        os.path.join(fe_input_dir, "custom_fields", "Custom_Fields.xlsx"),
        os.path.join(fe_input_dir, "Custom_Fields.xlsx"),
        os.path.join(main_input_dir, "custom_fields", "Custom_Fields.xlsx"),
        os.path.join(main_input_dir, "Custom_Fields.xlsx"),
    ]
    f2_paths = [
        os.path.join(fe_input_dir, "custom_fields", "Custom_Fields_Type_Menu.xlsx"),
        os.path.join(fe_input_dir, "Custom_Fields_Type_Menu.xlsx"),
        os.path.join(main_input_dir, "custom_fields", "Custom_Fields_Type_Menu.xlsx"),
        os.path.join(main_input_dir, "Custom_Fields_Type_Menu.xlsx"),
    ]

    f1 = next((p for p in f1_paths if os.path.exists(p)), None)
    f2 = next((p for p in f2_paths if os.path.exists(p)), None)

    cf_map = None
    if f1 or f2:
        cf_map = parse_custom_fields_excel(f1, f2)
        add_log(f"Loaded Custom Fields Excel mapping ({len(cf_map)} fields)", "INFO")

    add_log(f"Found {len(ws_files)} workspace XML(s) and {len(obj_files)} object XML(s)", "INFO")
    return _process_xml_files(ws_files, obj_files, uploaded_cf_map=cf_map)

@app.route("/api/fetch_rest_schemas", methods=["POST"])
def fetch_rest_schemas():
    """
    STRICT READ-ONLY: Fetches standard object schemas directly from OSVC Connect REST API
    using HTTP GET requests ONLY.
    """
    data = request.json or {}
    host = data.get("host", "").strip()
    username = data.get("username", "").strip()
    password = data.get("password", "").strip()
    include_custom = bool(data.get("include_custom", False))

    add_log("Received REST extraction request (STRICT HTTP GET ONLY)", "INFO")

    if not host or not username or not password:
        add_log("Missing required credentials (host/username/password)", "ERROR")
        return jsonify({
            "success": False,
            "error": "Host URL, Username, and Password are required."
        }), 400

    try:
        fetched_objects = fetch_standard_objects_via_rest(
            host=host,
            username=username,
            password=password,
            include_custom=include_custom,
            log_cb=add_log
        )

        if not fetched_objects:
            add_log("No standard object schemas returned from REST API.", "WARNING")
            return jsonify({
                "success": False,
                "error": "No standard object schemas could be extracted from the specified OSVC instance."
            }), 404

        existing_ws = RESULTS_CACHE.get("workspaces") or []
        if not existing_ws:
            sample_dir = os.path.join(CURRENT_DIR, "sample_inputs")
            if os.path.exists(sample_dir):
                for f in os.listdir(sample_dir):
                    if f.endswith(".xml") and "Workspace" in f:
                        try:
                            existing_ws.append(parse_workspace_xml(os.path.join(sample_dir, f)))
                        except Exception:
                            pass

        out_dir = os.path.join(CURRENT_DIR, "results")
        os.makedirs(out_dir, exist_ok=True)

        ws_xlsx_path = os.path.join(out_dir, "workspaces.xlsx")
        obj_xlsx_path = os.path.join(out_dir, "objects.xlsx")
        comb_xlsx_path = os.path.join(out_dir, "field_catalog.xlsx")

        write_workspaces_excel(existing_ws, fetched_objects, ws_xlsx_path)
        write_objects_excel(fetched_objects, obj_xlsx_path)
        write_combined_excel(existing_ws, fetched_objects, comb_xlsx_path)

        RESULTS_CACHE["standard_objects_map"] = fetched_objects
        _save_std_cache(fetched_objects)
        custom_map = RESULTS_CACHE.get("custom_objects_map") or {}
        combined_map = merge_objects_maps(fetched_objects, custom_map)
        RESULTS_CACHE["combined_objects_map"] = combined_map

        out_dir = os.path.join(CURRENT_DIR, "results")
        os.makedirs(out_dir, exist_ok=True)

        ws_xlsx_path            = os.path.join(out_dir, "Workspaces.xlsx")
        ws_no_ignored_xlsx_path = os.path.join(out_dir, "Workspaces_No_Ignored.xlsx")
        std_xlsx_path           = os.path.join(out_dir, "Standard_Objects.xlsx")
        cst_xlsx_path           = os.path.join(out_dir, "Custom_Objects.xlsx")
        comb_xlsx_path          = os.path.join(out_dir, "Field_Catalog.xlsx")
        comb_no_ignored_xlsx_path= os.path.join(out_dir, "Field_Catalog_No_Ignored.xlsx")
        cf_xlsx_path            = os.path.join(out_dir, "Custom_Fields_Mapping.xlsx")

        f1 = os.path.join(CURRENT_DIR, "..", "input", "custom_fields", "Custom_Fields.xlsx")
        f2 = os.path.join(CURRENT_DIR, "..", "input", "custom_fields", "Custom_Fields_Type_Menu.xlsx")
        cf_map = parse_custom_fields_excel(f1, f2) if (os.path.exists(f1) or os.path.exists(f2)) else {}

        write_objects_excel(fetched_objects, std_xlsx_path)
        if custom_map:
            write_objects_excel(custom_map, cst_xlsx_path)

        if cf_map:
            write_custom_fields_mapping_excel(cf_map, cf_xlsx_path)

        write_workspaces_excel(existing_ws, combined_map, ws_xlsx_path, include_ignored_tab=True, custom_fields_map=cf_map)
        write_workspaces_excel(existing_ws, combined_map, ws_no_ignored_xlsx_path, include_ignored_tab=False, custom_fields_map=cf_map)

        write_combined_excel(existing_ws, combined_map, comb_xlsx_path, include_ignored_tab=True, custom_fields_map=cf_map)
        write_combined_excel(existing_ws, combined_map, comb_no_ignored_xlsx_path, include_ignored_tab=False, custom_fields_map=cf_map)

        RESULTS_CACHE["workspaces"] = existing_ws
        RESULTS_CACHE["output_dir"] = out_dir

        add_log("Generated Excel workbooks: Standard_Objects.xlsx, Custom_Objects.xlsx, Custom_Fields_Mapping.xlsx, Workspaces.xlsx, Field_Catalog.xlsx", "SUCCESS")

        # Build JSON preview for UI (separate Standard Objects and Custom Objects)
        def _build_obj_preview(objs_dict):
            lst = []
            for o_name, o_data in (objs_dict or {}).items():
                disp_name = o_data.get("object_name", o_name)
                rows = []
                for of in o_data.get("fields", []):
                    rows.append({
                        "field_key": _obj_field_key(of),
                        "field_label": of.get("field_label", ""),
                        "data_type": of.get("data_type", ""),
                        "field_type": _field_type_from_obj(of),
                        "is_nullable": "Yes" if of.get("is_nullable") else "No",
                        "is_lookup": "Yes" if of.get("is_lookup") else "No",
                        "is_list": "Yes" if of.get("is_list") else "No",
                        "is_autoupdate": "Yes" if of.get("is_autoupdate") else "No",
                        "is_readonly": "Yes" if of.get("is_readonly") else "No",
                        "max_length": str(of.get("max_length", "-")),
                        "description": of.get("description", "")
                    })
                lst.append({
                    "object_name": disp_name,
                    "field_count": len(rows),
                    "rows": rows
                })
            return lst

        std_map_all = fetched_objects or RESULTS_CACHE.get("standard_objects_map") or {}
        cst_map_all = custom_map or RESULTS_CACHE.get("custom_objects_map") or {}

        preview_std_objects = _build_obj_preview(std_map_all)
        preview_cst_objects = _build_obj_preview(cst_map_all)

        preview_workspaces = _build_workspaces_preview(existing_ws, combined_map, custom_fields_map=cf_map)
        preview_combined   = _build_combined_preview(existing_ws, combined_map, custom_fields_map=cf_map)

        summary = {
            "workspace_count": len(existing_ws),
            "std_object_count": len(std_map_all),
            "cst_object_count": len(cst_map_all),
            "object_count": len(combined_map),
            "total_workspace_fields": sum(len(w.get("fields", [])) for w in existing_ws),
            "total_object_fields": sum(len(o.get("fields", [])) for o in combined_map.values()),
            "skipped_files": 0
        }
        RESULTS_CACHE["summary"] = summary

        return jsonify({
            "success": True,
            "summary": summary,
            "workspaces": preview_workspaces,
            "std_objects": preview_std_objects,
            "cst_objects": preview_cst_objects,
            "objects": preview_std_objects + preview_cst_objects,
            "combined": preview_combined
        })

    except Exception as err:
        tb = traceback.format_exc()
        add_log(f"REST API Connection error: {str(err)}", "ERROR")
        add_log(tb, "TRACEBACK")
        return jsonify({
            "success": False,
            "error": f"REST API Connection error: {str(err)}",
            "traceback": tb
        }), 200

@app.route("/api/extract", methods=["POST"])
def extract_files():
    """Handles file uploads for Workspace XMLs, Object XMLs, and Custom Field Excel files."""
    uploaded_ws_files = request.files.getlist("workspace_files")
    uploaded_obj_files = request.files.getlist("object_files")
    uploaded_cf_files = request.files.getlist("custom_fields_files")

    if not uploaded_ws_files and not uploaded_obj_files and not uploaded_cf_files:
        return jsonify({"success": False, "error": "No files uploaded."}), 400

    temp_dir = tempfile.mkdtemp(prefix="field_extractor_upload_")
    ws_temp_paths = []
    obj_temp_paths = []
    cf_temp_paths = []

    for f in uploaded_ws_files:
        if f and f.filename and f.filename.endswith(".xml"):
            s_name = secure_filename(f.filename) or "workspace.xml"
            dest = os.path.join(temp_dir, f"ws_{s_name}")
            f.save(dest)
            ws_temp_paths.append(dest)

    for f in uploaded_obj_files:
        if f and f.filename and f.filename.endswith(".xml"):
            s_name = secure_filename(f.filename) or "object.xml"
            dest = os.path.join(temp_dir, f"obj_{s_name}")
            f.save(dest)
            obj_temp_paths.append(dest)

    for f in uploaded_cf_files:
        if f and f.filename and (f.filename.endswith(".xlsx") or f.filename.endswith(".xls")):
            s_name = secure_filename(f.filename) or "custom_fields.xlsx"
            dest = os.path.join(temp_dir, f"cf_{s_name}")
            f.save(dest)
            cf_temp_paths.append(dest)

    add_log(f"Uploaded {len(ws_temp_paths)} workspace XML(s), {len(obj_temp_paths)} object XML(s), and {len(cf_temp_paths)} custom field Excel file(s)", "INFO")

    f1 = None
    f2 = None
    for p in cf_temp_paths:
        if "menu" in p.lower():
            f2 = p
        else:
            f1 = p

    cf_map = parse_custom_fields_excel(f1, f2) if (f1 or f2) else None
    return _process_xml_files(ws_temp_paths, obj_temp_paths, uploaded_cf_map=cf_map)


def _process_xml_files(ws_file_paths, obj_file_paths, auto_fetch_rest=False, uploaded_cf_map=None):
    """Processes workspace and object XML file paths, generates Excel files, and returns preview data."""
    parsed_objects = {}
    for o_path in obj_file_paths:
        try:
            res = parse_object_xml(o_path)
            o_items = res if isinstance(res, list) else ([res] if res else [])
            for o_data in o_items:
                obj_name = o_data.get("object_name", "")
                if obj_name:
                    parsed_objects[obj_name.lower()] = o_data
                    add_log(f"Parsed Object XML: {obj_name} ({len(o_data.get('fields', []))} fields)", "INFO")
        except Exception as e:
            add_log(f"Skipped non-valid Object XML: {os.path.basename(o_path)}", "WARNING")

    parsed_workspaces = []
    skipped_count = 0
    for w_path in ws_file_paths:
        try:
            w_data = parse_workspace_xml(w_path)
            parsed_workspaces.append(w_data)
            add_log(f"Parsed Workspace XML: {w_data['workspace_name']} ({len(w_data.get('fields', []))} fields)", "INFO")
        except Exception as e:
            skipped_count += 1
            add_log(f"Skipped non-valid Workspace XML: {os.path.basename(w_path)}", "WARNING")

    if not parsed_workspaces and not parsed_objects:
        add_log("No valid Workspace or Object XML files parsed.", "ERROR")
        return jsonify({
            "success": False,
            "error": "Failed to parse any valid Workspace or Object XML files."
        }), 400
    out_dir = os.path.join(CURRENT_DIR, "results")
    os.makedirs(out_dir, exist_ok=True)

    std_map = RESULTS_CACHE.get("standard_objects_map") or load_standard_objects_from_excel() or {}

    if not std_map and auto_fetch_rest:
        host_to_use = _cfg_host or BASE_URL
        user_to_use = _cfg_user or USERNAME
        pass_to_use = _cfg_pass or PASSWORD
        if host_to_use and user_to_use and pass_to_use:
            add_log("Auto-fetching ALL standard object schemas from OSVC Connect REST API...", "INFO")
            try:
                std_map = fetch_standard_objects_via_rest(
                    host=host_to_use,
                    username=user_to_use,
                    password=pass_to_use,
                    include_custom=True,
                    log_cb=add_log
                )
            except Exception as err:
                add_log(f"Auto-fetch REST schemas warning: {err}", "WARNING")

    if std_map:
        RESULTS_CACHE["standard_objects_map"] = std_map

    combined_map = merge_objects_maps(std_map, parsed_objects)
    RESULTS_CACHE["combined_objects_map"] = combined_map

    std_xlsx_path            = os.path.join(out_dir, "Standard_Objects.xlsx")
    cst_xlsx_path            = os.path.join(out_dir, "Custom_Objects.xlsx")
    ws_xlsx_path             = os.path.join(out_dir, "Workspaces.xlsx")
    ws_no_ignored_xlsx_path  = os.path.join(out_dir, "Workspaces_No_Ignored.xlsx")
    ws_simplified_xlsx_path  = os.path.join(out_dir, "Workspaces_Simplified.xlsx")
    comb_xlsx_path           = os.path.join(out_dir, "Field_Catalog.xlsx")
    comb_no_ignored_xlsx_path= os.path.join(out_dir, "Field_Catalog_No_Ignored.xlsx")
    comb_simplified_xlsx_path= os.path.join(out_dir, "Field_Catalog_Simplified.xlsx")
    cf_xlsx_path             = os.path.join(out_dir, "Custom_Fields_Mapping.xlsx")

    f1 = os.path.join(CURRENT_DIR, "..", "input", "custom_fields", "Custom_Fields.xlsx")
    f2 = os.path.join(CURRENT_DIR, "..", "input", "custom_fields", "Custom_Fields_Type_Menu.xlsx")
    cf_map = uploaded_cf_map or (parse_custom_fields_excel(f1, f2) if (os.path.exists(f1) or os.path.exists(f2)) else {})

    if parsed_objects:
        from field_extractor.parsers.custom_field_excel_parser import enrich_custom_fields_with_custom_objects
        cf_map = enrich_custom_fields_with_custom_objects(cf_map, parsed_objects)

    if std_map:
        write_objects_excel(std_map, std_xlsx_path, custom_fields_map=cf_map)
    write_objects_excel(parsed_objects, cst_xlsx_path, custom_fields_map=cf_map)

    if cf_map:
        write_custom_fields_mapping_excel(cf_map, cf_xlsx_path)

    write_workspaces_excel(parsed_workspaces, combined_map, ws_xlsx_path, include_ignored_tab=True, custom_fields_map=cf_map)
    write_workspaces_excel(parsed_workspaces, combined_map, ws_no_ignored_xlsx_path, include_ignored_tab=False, custom_fields_map=cf_map)
    write_workspaces_excel(parsed_workspaces, combined_map, ws_simplified_xlsx_path, include_ignored_tab=False, custom_fields_map=cf_map, simplify_attributes=True)

    write_combined_excel(parsed_workspaces, combined_map, comb_xlsx_path, include_ignored_tab=True, custom_fields_map=cf_map)
    write_combined_excel(parsed_workspaces, combined_map, comb_no_ignored_xlsx_path, include_ignored_tab=False, custom_fields_map=cf_map)
    write_combined_excel(parsed_workspaces, combined_map, comb_simplified_xlsx_path, include_ignored_tab=False, custom_fields_map=cf_map, simplify_attributes=True)

    def _build_obj_preview(objs_dict):
        lst = []
        for o_name, o_data in (objs_dict or {}).items():
            disp_name = o_data.get("object_name", o_name)
            rows = []
            for of in o_data.get("fields", []):
                # isEnumerable is REST-only; check key presence
                if "isEnumerable" not in of:
                    is_enum_str = "-"
                else:
                    is_enum_val = of["isEnumerable"]
                    if isinstance(is_enum_val, bool):
                        is_enum_str = "Yes" if is_enum_val else "No"
                    else:
                        is_enum_str = str(is_enum_val) if is_enum_val is not None else "-"

                items_val = of.get("items")
                if "items" not in of and not of.get("is_list"):
                    items_str = "-"
                elif isinstance(items_val, (dict, list)):
                    items_str = str(items_val)
                elif of.get("is_list"):
                    items_str = "IsList: Yes"
                else:
                    items_str = str(items_val) if items_val is not None else "-"

                ref_val = of.get("$ref") or of.get("ref") or of.get("ref_url") or "-"

                rows.append({
                    "field_key": _obj_field_key(of),
                    "field_label": of.get("field_label", ""),
                    "data_type": of.get("data_type", ""),
                    "field_type": _field_type_from_obj(of),
                    "is_system_field": "Yes" if of.get("is_system_field") else "No",
                    "package_name": of.get("package_name", ""),
                    "is_nullable": "Yes" if of.get("is_nullable") else "No",
                    "is_lookup": "Yes" if of.get("is_lookup") else "No",
                    "is_readonly": "Yes" if of.get("is_readonly") else "No",
                    "max_length": str(of.get("max_length", "-")),
                    "description": of.get("description", ""),
                    "is_available_get":   ("-" if "is_available_get"   not in of else "Yes" if of["is_available_get"]   else "No"),
                    "is_available_post":  ("-" if "is_available_post"  not in of else "Yes" if of["is_available_post"]  else "No"),
                    "is_available_patch": ("-" if "is_available_patch" not in of else "Yes" if of["is_available_patch"] else "No"),
                    "is_deprecated":      ("-" if "is_deprecated"      not in of else "Yes" if of["is_deprecated"]      else "No"),
                    "is_enumerable": is_enum_str,
                    "minimum": "-" if "minimum" not in of else str(of["minimum"]),
                    "maximum": "-" if "maximum" not in of else str(of["maximum"]),
                    "ref": str(ref_val),
                    "items": items_str,
                    "pattern": str(of.get("pattern", "-"))
                })
            lst.append({
                "object_name": disp_name,
                "field_count": len(rows),
                "rows": rows
            })
        return lst

    RESULTS_CACHE["workspaces"] = parsed_workspaces
    RESULTS_CACHE["output_dir"] = out_dir

    add_log("Generated Excel files: Standard_Objects.xlsx, Custom_Objects.xlsx, Workspaces.xlsx, Field_Catalog.xlsx", "SUCCESS")

    # Build JSON preview payload for UI (1-to-1 match with generated Excel files)
    preview_std_objects = _build_obj_preview(std_map)
    preview_cst_objects = _build_obj_preview(parsed_objects)

    preview_workspaces = _build_workspaces_preview(parsed_workspaces, combined_map, custom_fields_map=cf_map)
    preview_combined   = _build_combined_preview(parsed_workspaces, combined_map, custom_fields_map=cf_map)

    summary = {
        "workspace_count": len(parsed_workspaces),
        "std_object_count": len(std_map),
        "cst_object_count": len(parsed_objects),
        "object_count": len(combined_map),
        "total_workspace_fields": sum(len(w.get("fields", [])) for w in parsed_workspaces),
        "total_object_fields": sum(len(o.get("fields", [])) for o in combined_map.values()),
        "skipped_files": skipped_count
    }
    RESULTS_CACHE["summary"] = summary

    add_log("Processing completed successfully.", "SUCCESS")

    return jsonify({
        "success": True,
        "summary": summary,
        "workspaces": preview_workspaces,
        "std_objects": preview_std_objects,
        "cst_objects": preview_cst_objects,
        "objects": preview_std_objects + preview_cst_objects,
        "combined": preview_combined
    })

@app.route("/api/download/<filename>")
def download_file(filename):
    """Downloads generated Excel files or ZIP package."""
    out_dir = RESULTS_CACHE.get("output_dir") or os.path.join(CURRENT_DIR, "results")

    valid_files = [
        "Standard_Objects.xlsx", "Custom_Objects.xlsx", "Workspaces.xlsx", "Field_Catalog.xlsx",
        "standard_objects.xlsx", "custom_objects.xlsx", "workspaces.xlsx", "field_catalog.xlsx", "combined.xlsx"
    ]
    if filename in valid_files:
        file_path = os.path.join(out_dir, filename)
        if not os.path.exists(file_path):
            # Try case-insensitive / legacy match
            for existing in os.listdir(out_dir) if os.path.exists(out_dir) else []:
                if existing.lower() == filename.lower():
                    file_path = os.path.join(out_dir, existing)
                    break
        if os.path.exists(file_path):
            return send_file(file_path, as_attachment=True)
        return jsonify({"error": f"File {filename} not found."}), 404

    if filename == "all_reports.zip":
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as z:
            for fname in valid_files:
                fpath = os.path.join(out_dir, fname)
                if os.path.exists(fpath):
                    z.write(fpath, arcname=fname)
        zip_buffer.seek(0)
        return send_file(
            zip_buffer,
            mimetype="application/zip",
            as_attachment=True,
            download_name="OSVC_Field_Extractor_Excel_Reports.zip"
        )

    return jsonify({"error": "Invalid file request."}), 400


def run_server(port=5050, debug=False):
    add_log(f"Field Extractor Web UI Server running on port {port}", "INFO")
    print("==========================================================================")
    print(f"       OSVC FIELD EXTRACTOR WEB UI SERVER RUNNING ON PORT {port}")
    print(f"       Open browser to: http://localhost:{port}")
    print("==========================================================================")
    app.run(host="0.0.0.0", port=port, debug=debug)

if __name__ == "__main__":
    port_arg = 5050
    if len(sys.argv) > 1:
        try:
            port_arg = int(sys.argv[1])
        except ValueError:
            pass
    run_server(port=port_arg)

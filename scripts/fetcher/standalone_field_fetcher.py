#!/usr/bin/env python3
"""
Standalone OSVC Field Fetcher & Excel Generator Script.

Connects to Oracle Service Cloud (OSVC) Connect REST API, extracts standard & custom
object metadata schemas, and exports the result to a multi-tab Excel (.xlsx) file.

Usage:
    python scripts/fetcher/standalone_field_fetcher.py --host <HOST> --username <USER> --password <PASS> --output results/Fetched_Fields.xlsx
"""

import os
import sys
import argparse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure current script directory and parent directories are in python module search path
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
ROOT_DIR = os.path.dirname(PARENT_DIR)

for d in (SCRIPT_DIR, PARENT_DIR, ROOT_DIR):
    if d and d not in sys.path:
        sys.path.insert(0, d)

try:
    from osvc_rest_fetcher import fetch_standard_objects_via_rest, KNOWN_STANDARD_OBJECTS
except ImportError:
    try:
        from fetcher.osvc_rest_fetcher import fetch_standard_objects_via_rest, KNOWN_STANDARD_OBJECTS
    except ImportError:
        from scripts.fetcher.osvc_rest_fetcher import fetch_standard_objects_via_rest, KNOWN_STANDARD_OBJECTS

# Try to load connection defaults from field_extractor/config.py if available
DEFAULT_HOST = os.environ.get("OSVC_HOST", "")
DEFAULT_USER = os.environ.get("OSVC_USERNAME", "")
DEFAULT_PASS = os.environ.get("OSVC_PASSWORD", "")

if not DEFAULT_HOST or not DEFAULT_USER or not DEFAULT_PASS:
    try:
        config_path = os.path.join(ROOT_DIR, "field_extractor", "config.py")
        if os.path.exists(config_path):
            import importlib.util
            spec = importlib.util.spec_from_file_location("field_extractor_config", config_path)
            cfg_mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(cfg_mod)
            DEFAULT_HOST = getattr(cfg_mod, "BASE_URL", DEFAULT_HOST)
            DEFAULT_USER = getattr(cfg_mod, "USERNAME", DEFAULT_USER)
            DEFAULT_PASS = getattr(cfg_mod, "PASSWORD", DEFAULT_PASS)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Excel Styling Definitions
# ---------------------------------------------------------------------------

HEADER_FILL   = PatternFill("solid", fgColor="2E75B6")   # Professional blue
SUMMARY_FILL  = PatternFill("solid", fgColor="1F4E78")   # Dark blue header
ALT_ROW_FILL  = PatternFill("solid", fgColor="F2F7FA")   # Soft light blue zebra
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT    = Font(name="Calibri", bold=True, color="1F4E78", size=14)
SUBTITLE_FONT = Font(name="Calibri", italic=True, color="595959", size=10)
BODY_FONT     = Font(name="Calibri", size=10)
BOLD_FONT     = Font(name="Calibri", bold=True, size=10)

THIN_BORDER   = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9")
)

HEADERS = [
    "Object Name",
    "Field ID",
    "Field Name",
    "Field Label",
    "Data Type",
    "Is System Field",
    "Package Name",
    "Is Nullable",
    "Is Lookup",
    "Is Read Only",
    "Max Length",
    "Description",
    "GET Available",
    "POST Available",
    "PATCH Available",
    "Is Deprecated",
    "$ref Target"
]


def _style_header_row(ws, headers, fill=HEADER_FILL):
    """Styles the header row of a worksheet with frozen panes and auto-filters."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"
    ws.row_dimensions[1].height = 28


def _auto_adjust_column_widths(ws, max_cols=None):
    """Adjusts column widths based on maximum string length per column."""
    num_cols = max_cols or ws.max_column
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)


def _safe_sheet_name(name):
    """Sanitizes string for a valid Excel sheet tab name (max 31 chars)."""
    clean_name = "".join([c if c.isalnum() or c in (" ", "_") else "_" for c in name])
    return clean_name[:31]


def export_to_excel(objects_map, output_filepath, host_url=""):
    """
    Exports the fetched objects_map into a multi-tab Excel workbook (.xlsx).
    Tabs generated:
    1. Summary - Metadata summary & per-object metrics
    2. All_Fields - Consolidated table of all fields across all objects
    3...N - Individual sheets for each extracted object
    """
    os.makedirs(os.path.dirname(os.path.abspath(output_filepath)), exist_ok=True)
    wb = openpyxl.Workbook()

    # ---------------------------------------------------------------------------
    # Sheet 1: Summary Page
    # ---------------------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    ws_summary.cell(row=1, column=1, value="OSVC REST API Metadata Extraction Summary").font = TITLE_FONT
    ws_summary.cell(row=2, column=1, value=f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = SUBTITLE_FONT

    metadata = [
        ("Host Endpoint URL", host_url or "N/A"),
        ("Total Objects Extracted", len(objects_map)),
        ("Total Fields Extracted", sum(len(o.get("fields", [])) for o in objects_map.values())),
        ("Total System Fields", sum(len([f for f in o.get("fields", []) if f.get("is_system_field")]) for o in objects_map.values())),
        ("Total Custom Fields", sum(len([f for f in o.get("fields", []) if not f.get("is_system_field")]) for o in objects_map.values()))
    ]

    ws_summary.cell(row=4, column=1, value="Metric").font = HEADER_FONT
    ws_summary.cell(row=4, column=1).fill = SUMMARY_FILL
    ws_summary.cell(row=4, column=2, value="Value").font = HEADER_FONT
    ws_summary.cell(row=4, column=2).fill = SUMMARY_FILL

    for idx, (m_name, m_val) in enumerate(metadata, start=5):
        cell_a = ws_summary.cell(row=idx, column=1, value=m_name)
        cell_b = ws_summary.cell(row=idx, column=2, value=m_val)
        cell_a.font = BOLD_FONT
        cell_b.font = BODY_FONT
        cell_a.border = THIN_BORDER
        cell_b.border = THIN_BORDER

    # Per-Object Metrics Table on Summary
    ws_summary.cell(row=12, column=1, value="Object Breakdown").font = TITLE_FONT

    breakdown_headers = ["Object Name", "Total Fields", "System Fields", "Custom Fields", "Lookup Fields"]
    for c_idx, bh in enumerate(breakdown_headers, start=1):
        cell = ws_summary.cell(row=14, column=c_idx, value=bh)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    row_pos = 15
    for obj_key, obj_data in sorted(objects_map.items()):
        o_name = obj_data.get("object_name", obj_key)
        fields = obj_data.get("fields", [])
        tot_f = len(fields)
        sys_f = len([f for f in fields if f.get("is_system_field")])
        cst_f = tot_f - sys_f
        lkp_f = len([f for f in fields if f.get("is_lookup")])

        row_vals = [o_name, tot_f, sys_f, cst_f, lkp_f]
        fill = ALT_ROW_FILL if row_pos % 2 == 0 else None
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws_summary.cell(row=row_pos, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if c_idx > 1:
                cell.alignment = Alignment(horizontal="center")
        row_pos += 1

    _auto_adjust_column_widths(ws_summary, max_cols=5)

    # ---------------------------------------------------------------------------
    # Sheet 2: All_Fields Consolidated Catalog
    # ---------------------------------------------------------------------------
    ws_all = wb.create_sheet(title="All_Fields")
    ws_all.views.sheetView[0].showGridLines = True
    _style_header_row(ws_all, HEADERS)

    all_row_idx = 2
    for obj_key, obj_data in sorted(objects_map.items()):
        obj_disp_name = obj_data.get("object_name", obj_key.capitalize())
        fields = obj_data.get("fields", [])

        for field in fields:
            row_data = [
                obj_disp_name,
                field.get("field_id", "-"),
                field.get("field_name", "-"),
                field.get("field_label", "-"),
                field.get("data_type", "-"),
                "Yes" if field.get("is_system_field") else "No",
                field.get("package_name", "-"),
                "Yes" if field.get("is_nullable") else "No",
                "Yes" if field.get("is_lookup") else "No",
                "Yes" if field.get("is_readonly") else "No",
                str(field.get("max_length", "-")),
                field.get("description", "-"),
                "Yes" if field.get("is_available_get") else "No",
                "Yes" if field.get("is_available_post") else "No",
                "Yes" if field.get("is_available_patch") else "No",
                "Yes" if field.get("is_deprecated") else "No",
                str(field.get("$ref", "-"))
            ]

            fill = ALT_ROW_FILL if all_row_idx % 2 == 0 else None
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws_all.cell(row=all_row_idx, column=c_idx, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                if fill:
                    cell.fill = fill
            all_row_idx += 1

    _auto_adjust_column_widths(ws_all)

    # ---------------------------------------------------------------------------
    # Sheets 3...N: Per-Object Worksheets
    # ---------------------------------------------------------------------------
    for obj_key, obj_data in sorted(objects_map.items()):
        obj_disp_name = obj_data.get("object_name", obj_key.capitalize())
        sheet_title = _safe_sheet_name(obj_disp_name)

        ws_obj = wb.create_sheet(title=sheet_title)
        ws_obj.views.sheetView[0].showGridLines = True
        _style_header_row(ws_obj, HEADERS[1:])  # Skip 'Object Name' column for individual tab

        fields = obj_data.get("fields", [])
        for f_idx, field in enumerate(fields, start=2):
            row_data = [
                field.get("field_id", "-"),
                field.get("field_name", "-"),
                field.get("field_label", "-"),
                field.get("data_type", "-"),
                "Yes" if field.get("is_system_field") else "No",
                field.get("package_name", "-"),
                "Yes" if field.get("is_nullable") else "No",
                "Yes" if field.get("is_lookup") else "No",
                "Yes" if field.get("is_readonly") else "No",
                str(field.get("max_length", "-")),
                field.get("description", "-"),
                "Yes" if field.get("is_available_get") else "No",
                "Yes" if field.get("is_available_post") else "No",
                "Yes" if field.get("is_available_patch") else "No",
                "Yes" if field.get("is_deprecated") else "No",
                str(field.get("$ref", "-"))
            ]

            fill = ALT_ROW_FILL if f_idx % 2 == 0 else None
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws_obj.cell(row=f_idx, column=c_idx, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                if fill:
                    cell.fill = fill

        _auto_adjust_column_widths(ws_obj)

    wb.save(output_filepath)
    print(f"[SUCCESS] Excel workbook successfully generated: {output_filepath}")


DEFAULT_OUTPUT_PATH = os.path.join(SCRIPT_DIR, "results", "Fetched_Fields.xlsx")

def main():
    parser = argparse.ArgumentParser(description="Standalone OSVC REST Field Fetcher & Excel Generator")
    parser.add_argument("--host", default=DEFAULT_HOST, help="OSVC Host domain or full endpoint URL")
    parser.add_argument("--username", default=DEFAULT_USER, help="OSVC REST API Username")
    parser.add_argument("--password", default=DEFAULT_PASS, help="OSVC REST API Password")
    parser.add_argument("--output", default=DEFAULT_OUTPUT_PATH, help="Output Excel file path (.xlsx)")
    parser.add_argument("--objects", nargs="+", default=None, help="List of standard objects to fetch (e.g. contacts incidents organizations)")
    parser.add_argument("--include-custom", action="store_true", help="Include custom objects / custom fields in extraction")

    args = parser.parse_args()

    if not args.host or not args.username or not args.password:
        print("[ERROR] Missing OSVC connection credentials.")
        print("Please provide --host, --username, and --password as arguments or set environment variables OSVC_HOST, OSVC_USERNAME, OSVC_PASSWORD.")
        sys.exit(1)

    print("[START] Connecting to OSVC Connect REST API...")
    print(f"[INFO] Target Host: {args.host}")
    print(f"[INFO] Username   : {args.username}")
    print(f"[INFO] Output Path: {args.output}")

    if args.objects:
        print(f"[INFO] Filtering for selected objects: {', '.join(args.objects)}")
    else:
        print(f"[INFO] Target Objects: Default standard objects list ({len(KNOWN_STANDARD_OBJECTS)} objects)")

    try:
        objects_map = fetch_standard_objects_via_rest(
            host=args.host,
            username=args.username,
            password=args.password,
            selected_objects=args.objects,
            include_custom=args.include_custom,
            log_cb=print
        )

        if not objects_map:
            print("[WARNING] No object metadata schemas were fetched from the API.")
            sys.exit(1)

        tot_fields = sum(len(o.get("fields", [])) for o in objects_map.values())
        print(f"[SUCCESS] Fetched metadata for {len(objects_map)} object(s) containing {tot_fields} total field definitions.")

        print("[INFO] Generating Excel report...")
        export_to_excel(objects_map, args.output, host_url=args.host)
        print(f"[DONE] Extraction completed successfully. Excel saved at: {args.output}")

    except Exception as err:
        print(f"[ERROR] Extraction failed: {err}")
        sys.exit(1)


if __name__ == "__main__":
    main()

"""
Custom Field Excel Parser for OSVC Field Extractor.

Parses OSVC Custom Fields export Excel files:
1. Custom Fields export (grouped by Table: Contact, Ticket/Incident, Answer)
2. Custom Fields of Type Menu export (Name, Column Name, Menu Items)
"""

import os
import openpyxl

def parse_custom_fields_excel(custom_fields_file=None, custom_menu_fields_file=None):
    """
    Parses OSVC Custom Field Excel exports and returns a dict mapping column_name -> custom_field_dict.
    Key format: 'c$col_name' or 'table.c$col_name' e.g. 'c$meps', 'contact.c$contact_id'.
    """
    fields_map = {}

    # 1. Parse Menu Items from Custom Fields of Type Menu file
    menu_items_map = {}
    if custom_menu_fields_file and os.path.exists(custom_menu_fields_file):
        try:
            wb_menu = openpyxl.load_workbook(custom_menu_fields_file, data_only=True)
            ws_menu = wb_menu.active
            for idx, row in enumerate(ws_menu.iter_rows(values_only=True), start=1):
                if idx < 3:
                    continue  # Skip title & header
                name = str(row[0]).strip() if row[0] else ""
                col_name = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ""
                menu_str = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                if col_name and menu_str:
                    items_list = [m.strip() for m in menu_str.split(":") if m.strip()]
                    menu_items_map[col_name] = {
                        "name": name,
                        "menu_items": items_list,
                        "menu_items_str": menu_str
                    }
        except Exception as e:
            print(f"[WARNING] Could not parse Menu Custom Fields file: {e}")

    # 2. Parse Custom Fields export file
    if custom_fields_file and os.path.exists(custom_fields_file):
        try:
            wb = openpyxl.load_workbook(custom_fields_file, data_only=True)
            ws = wb.active

            current_table = "Incident"
            for row in ws.iter_rows(values_only=True):
                cell1 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                cell2 = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                cell3 = str(row[3]).strip() if len(row) > 3 and row[3] else ""

                if cell1.startswith("Table:"):
                    raw_tbl = cell1.replace("Table:", "").strip()
                    # Map OSVC internal table names: Ticket -> Incident
                    current_table = "Incident" if raw_tbl.lower() in ("ticket", "incident") else raw_tbl
                    continue

                if cell2 and cell3 and cell3.lower().startswith("c$"):
                    col_name = cell3.lower()
                    cf_id = row[4] if len(row) > 4 else None
                    dtype = str(row[5]).strip() if len(row) > 5 and row[5] else "Text"
                    fsize = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                    flags = str(row[10]).strip() if len(row) > 10 and row[10] else ""

                    menu_info = menu_items_map.get(col_name, {})

                    field_entry = {
                        "table_name": current_table,
                        "field_label": cell2,
                        "column_name": col_name,
                        "custom_field_id": cf_id,
                        "data_type": dtype,
                        "field_size": fsize,
                        "flags": flags,
                        "menu_items": menu_info.get("menu_items", []),
                        "menu_items_str": menu_info.get("menu_items_str", ""),
                        "is_custom_field": True
                    }

                    tbl_low   = current_table.lower()
                    col_bare  = col_name.replace("c$", "") if col_name.startswith("c$") else col_name
                    col_no_us = col_name.replace("_", "")

                    for k in (col_name, col_bare, col_no_us, f"{tbl_low}.{col_name}", f"{tbl_low}.{col_bare}", f"{tbl_low}.${col_name}", f"{tbl_low}.${col_bare}", f"c${col_bare}"):
                        if k:
                            fields_map[k] = field_entry

        except Exception as e:
            print(f"[WARNING] Could not parse Custom Fields file: {e}")

    # Also add any standalone menu fields from File 2 not in File 1
    for col_name, m_info in menu_items_map.items():
        if col_name not in fields_map:
            field_entry = {
                "table_name": "Incident",
                "field_label": m_info["name"],
                "column_name": col_name,
                "custom_field_id": None,
                "data_type": "Menu",
                "field_size": "",
                "flags": "",
                "menu_items": m_info["menu_items"],
                "menu_items_str": m_info["menu_items_str"],
                "is_custom_field": True
            }
            col_no_us = col_name.replace("_", "")
            for k in (col_name, col_no_us, f"incident.{col_name}", f"incident.{col_no_us}", f"incident.${col_name}", f"incident.${col_no_us}"):
                if k:
                    fields_map[k] = field_entry

    return fields_map


import re

def enrich_custom_fields_with_custom_objects(cf_map, custom_objects_map):
    """
    Enriches custom_fields_map with Global Custom Menu Objects (e.g. ContractType, CarrierName in package VSP).
    Extracts menu options from <MenuItems> in CustomObject XML files and attaches them to candidate keys.
    """
    if cf_map is None:
        cf_map = {}
    if not custom_objects_map:
        return cf_map

    for obj_key, o_data in custom_objects_map.items():
        menu_items = o_data.get("menu_items") or []
        if not menu_items:
            continue

        co_name = o_data.get("object_name", "")
        co_label = o_data.get("co_label") or co_name
        pkg_name = o_data.get("package_name") or "VSP"
        menu_str = ":".join(menu_items)

        # Generate candidates for mapping
        col_raw = f"{pkg_name.lower()}${co_name.lower()}"
        col_bare = co_name.lower()
        col_c = f"c${co_name.lower()}"

        rest_snake = re.sub(r'([a-z0-9])([A-Z])', r'\1_\2', co_name).lower()
        snake_bare = rest_snake
        snake_c = f"c${snake_bare}"
        snake_pkg = f"{pkg_name.lower()}${snake_bare}"

        field_entry = {
            "table_name": pkg_name,
            "field_label": co_label,
            "column_name": col_raw,
            "custom_field_id": o_data.get("id", "N/A"),
            "data_type": "Menu",
            "field_size": "—",
            "flags": f"Global Custom Menu Object ({pkg_name})",
            "menu_items": menu_items,
            "menu_items_str": menu_str,
            "is_custom_field": True
        }

        candidates = [
            col_raw, col_bare, col_c, snake_bare, snake_c, snake_pkg,
            f"{pkg_name.lower()}.${co_name.lower()}",
            f"{pkg_name.lower()}.${snake_bare}",
            f"incident.{col_raw}", f"incident.{col_c}", f"incident.{snake_c}",
            f"contact.{col_raw}", f"contact.{col_c}", f"contact.{snake_c}"
        ]

        for cand in candidates:
            if cand in cf_map:
                cf_map[cand]["menu_items"] = menu_items
                cf_map[cand]["menu_items_str"] = menu_str
                if cf_map[cand].get("data_type") in ("Text", "Integer", "—", ""):
                    cf_map[cand]["data_type"] = "Menu"
            else:
                cf_map[cand] = field_entry

    return cf_map

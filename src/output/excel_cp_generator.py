"""
Excel Customer Portal Generator for Config Accelerator.

Generates DTMO Customer Portal Implementation Excel workbook with exact 7 sheets:
- Summary
- Models
- Hooks
- Templates
- Pages
- Widgets
- Community
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADER_FILL  = PatternFill('solid', fgColor='2E75B6')
HEADER_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
BODY_FONT    = Font(name='Calibri', size=10)
THIN_BORDER  = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)


def _style_sheet(sheet, headers):
    sheet.views.sheetView[0].showGridLines = True
    sheet.row_dimensions[1].height = 24
    for col_idx, h in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _write_rows(sheet, rows):
    for r_idx, row in enumerate(rows, start=2):
        sheet.row_dimensions[r_idx].height = 20
        for col_idx, val in enumerate(row, start=1):
            cell = sheet.cell(row=r_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    # Auto-adjust column widths
    for col in sheet.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            max_len = max(max_len, len(val_str))
        sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 15), 80)


def generate_cp_excel_report(cp_data, output_path):
    """
    Generates the Customer Portal Implementation Excel workbook.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # 1. Summary Sheet
    ws_sum = wb.create_sheet(title="Summary")
    headers_sum = ['Point', 'Description', 'Comments']
    _style_sheet(ws_sum, headers_sum)
    rows_sum = [[item['point'], item['description'], item['comments']] for item in cp_data.get("summary", [])]
    _write_rows(ws_sum, rows_sum)

    # 2. Models Sheet
    ws_mod = wb.create_sheet(title="Models")
    headers_mod = ['Name', 'Purpose', 'Called By / Used In']
    _style_sheet(ws_mod, headers_mod)
    rows_mod = [[item['name'], item['purpose'], item['called_by']] for item in cp_data.get("models", [])]
    _write_rows(ws_mod, rows_mod)

    # 3. Hooks Sheet
    ws_hook = wb.create_sheet(title="Hooks")
    headers_hook = ['Name', 'Model', 'Purpose', 'Triggered When', 'Pages Affected']
    _style_sheet(ws_hook, headers_hook)
    rows_hook = [[item['name'], item['model'], item['purpose'], item['triggered_when'], item['pages_affected']] for item in cp_data.get("hooks", [])]
    _write_rows(ws_hook, rows_hook)

    # 4. Templates Sheet
    ws_tpl = wb.create_sheet(title="Templates")
    headers_tpl = ['Name', 'Purpose / Explanation', 'Used By']
    _style_sheet(ws_tpl, headers_tpl)
    rows_tpl = [[item['name'], item['purpose'], item['used_by']] for item in cp_data.get("templates", [])]
    _write_rows(ws_tpl, rows_tpl)

    # 5. Pages Sheet
    ws_pg = wb.create_sheet(title="Pages")
    headers_pg = ['Page / File', 'Key Widgets', 'Purpose', 'Login Required?']
    _style_sheet(ws_pg, headers_pg)
    rows_pg = [[item['page_file'], item['key_widgets'], item['purpose'], item['login_required']] for item in cp_data.get("pages", [])]
    _write_rows(ws_pg, rows_pg)

    # 6. Widgets Sheet
    ws_wdg = wb.create_sheet(title="Widgets")
    headers_wdg = ['Name', 'File / Path', 'Purpose / Explanation', 'Used In Pages']
    _style_sheet(ws_wdg, headers_wdg)
    rows_wdg = [[item['name'], item['file_path'], item['purpose'], item['used_in_pages_str']] for item in cp_data.get("widgets", [])]
    _write_rows(ws_wdg, rows_wdg)

    # 7. Community Sheet
    ws_com = wb.create_sheet(title="Community")
    headers_com = ['URL', 'Description']
    _style_sheet(ws_com, headers_com)
    rows_com = [[item['url'], item['description']] for item in cp_data.get("community", [])]
    _write_rows(ws_com, rows_com)

    out_dir = os.path.dirname(os.path.abspath(output_path))
    os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    return output_path

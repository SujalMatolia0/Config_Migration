"""
Config Accelerator Workspace Excel Exporter.

Generates a clean, integration-only multi-tab Excel workbook for OSVC Workspaces:
- Master Overview Sheet: Workspace_Overview
- Workspace sheets: Split into Table 1 (Layout Controls) and Table 2 (Workspace Rules).
- Table 2 (Workspace Rules) has a "Workspace Rules" title banner and 4 distinct columns: Rule Name, Trigger, Condition, Action.
- Dynamic Column Pruning: Automatically removes any column that has no non-zero / non-empty data across all rows.
"""

import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL   = PatternFill("solid", fgColor="2E75B6")   # Professional Blue
SECTION_FILL  = PatternFill("solid", fgColor="1F497D")   # Dark Navy Section Header
ALT_ROW_FILL  = PatternFill("solid", fgColor="F2F7FA")   # Light tint
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SECTION_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT     = Font(name="Calibri", size=10)
THIN_BORDER   = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

def _safe_tab_name(name):
    name = re.sub(r'[\\/*?:\[\]]', '_', str(name))
    return name[:31]

def _collect_all_tabs(tabs_list):
    flat = []
    for tab in tabs_list:
        flat.append(tab)
        for ts in tab.get("nested_tabsets", []):
            sub_tabs = ts.get("sub_tabs", [])
            flat.extend(_collect_all_tabs(sub_tabs))
    return flat

def _get_control_contents(xml_tag, field_code, acid=None, srid=None, extra=None):
    has_acid = (acid and str(acid) not in ("0", "None"))
    has_srid = (srid and str(srid) not in ("0", "None"))
    
    if xml_tag == "Report" or has_acid or has_srid:
        reports = []
        if has_acid: reports.append(f"AcId {acid}")
        if has_srid: reports.append(f"SearchId {srid}")
        if reports:
            return f"Report ({', '.join(reports)})"
        return "Report"
    
    if xml_tag == "Rule":
        return "Business Rule"
    if xml_tag == "AddInItem":
        return extra or "BUI / Desktop Add-In"

    fc = str(field_code or "").lower()
    if "attachment" in fc or "file" in fc:
        return "File Attachments"
    if "audit" in fc or "log" in fc:
        return "Audit Log"
    return "Sub-List Data"

def _format_action_item(a, rule_name=""):
    obj = a.get("object") or a.get("object_type") or "Unknown"
    obj_clean = str(obj).replace("Standard:", "")
    oper = a.get("operation") or ""
    val = a.get("value")
    
    if "DecisionTree" in obj_clean or obj == "DecisionTree":
        workflow_desc = rule_name if rule_name else "Guided Workflow"
        return f"Execute Guided Workflow: {workflow_desc}"
    
    oper_str = "Set Value" if oper == "Value" else oper
    val_str = str(val) if val is not None else ""
    
    parts = [f"Field: {obj_clean}"]
    if oper_str:
        parts.append(f"Operation: {oper_str}")
    if val_str:
        parts.append(f"Value: {val_str}")
        
    return " | ".join(parts)

def _get_rule_components(rule):
    """Formats business rules into separate (trigger_str, condition_str, action_str) strings."""
    trigs = rule.get("triggers", [])
    formatted_trigs = []
    for t in trigs:
        ft = (t.replace("FieldValueChanged", "On Field Change")
               .replace("EditorLoaded", "On Form Load")
               .replace("Saved", "On Form Save"))
        formatted_trigs.append(ft)

    trig_str = ", ".join(formatted_trigs) if formatted_trigs else "On Event"

    # Conditions (If)
    conds = rule.get("conditions", [])
    cond_parts = []
    for c in conds:
        c_src = c.get("source") or c.get("property") or ""
        c_op  = str(c.get("operator") or "=").replace("EQ", "==").replace("NE_OR_NULL", "!=").replace("NE", "!=")
        c_val = c.get("value") or ""
        if c_src or c_val:
            cond_parts.append(f"{c_src} {c_op} {c_val}".strip())
    
    cond_str = " AND ".join(cond_parts) if cond_parts else (rule.get("notes") or "—")

    # Actions
    rule_name = str(rule.get("name") or "").strip()
    formatted_actions = []
    
    for a in rule.get("actions", []):
        act_desc = _format_action_item(a, rule_name)
        if act_desc:
            formatted_actions.append(f"[{act_desc}]")

    act_str = "; ".join(formatted_actions) if formatted_actions else "Rule Action"

    return trig_str, cond_str, act_str

def _is_external_script(script):
    """Determines whether a Custom Script is Internal or External based on API calls and URL domain formats."""
    soap_cnt = len(script.get("external_soap_apis", []))
    rest_cnt = len(script.get("external_rest_apis", []))
    has_curl = script.get("has_curl", False)
    ext_calls = len(script.get("external_calls", []))

    if soap_cnt > 0 or rest_cnt > 0 or has_curl or ext_calls > 0:
        return True

    urls = script.get("urls", [])
    for u in urls:
        u_low = str(u).lower()
        if u_low.startswith("http://") or u_low.startswith("https://"):
            if not ("rnw" in u_low or "rightnow" in u_low or "custhelp" in u_low or "cgi-bin" in u_low):
                return True

    return False

def prune_empty_table(headers, rows):
    """
    Removes columns from headers and rows if every row has no data (None, empty string, 0, 'N/A')
    for that column.
    """
    if not rows:
        return [], []

    num_cols = len(headers)
    active_col_indices = []

    for c_idx in range(num_cols):
        has_data = False
        for row in rows:
            val = row[c_idx] if c_idx < len(row) else None
            if val is not None:
                if isinstance(val, (int, float)):
                    if val != 0:
                        has_data = True
                        break
                else:
                    s_val = str(val).strip()
                    if s_val not in ("", "0", "None", "N/A", "—"):
                        has_data = True
                        break
        if has_data:
            active_col_indices.append(c_idx)

    if not active_col_indices:
        return [], []

    pruned_headers = [headers[i] for i in active_col_indices]
    pruned_rows = []
    for row in rows:
        pruned_rows.append([row[i] if i < len(row) else None for i in active_col_indices])

    return pruned_headers, pruned_rows

def generate_workspaces_excel_report(workspaces, output_path, all_components=None, include_overview=True):
    """
    Generates a clean Integration-Only multi-tab Excel workbook from Config Accelerator parsed workspaces.
    - Master Overview Sheet: Workspace_Overview (Optional via include_overview)
    - Workspace sheets: Split into Table 1 (Layout Controls) and Table 2 (Workspace Rules).
    - Table 2: 4 distinct columns (Rule Name, Trigger, Condition, Action).
    - Completely suppresses Table 2 if a workspace has 0 rules configured.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    raw_overview_headers = [
        "Workspace Name", "Bound Entity", "Total Tabs", "Total Integrations",
        "Analytics Reports", "Relationship Items", "BUI Add-Ins",
        "Workspace & Object Rules", "Custom Scripts - Internal", "Custom Scripts - External"
    ]

    overview_rows = []

    # Extra components from analysis pipeline
    all_components = all_components or {}
    script_list = all_components.get("scripts") or all_components.get("customScripts") or []
    rule_list   = all_components.get("businessRules") or all_components.get("rules") or []

    for ws in sorted(workspaces, key=lambda x: str(x.get("name", ""))):
        ws_name   = ws.get("name") or ws.get("workspace_name") or "Workspace"
        bound_obj = ws.get("type") or ws.get("ui_type") or ws.get("bound_object") or "Incident"
        
        ctrl_items = []
        rule_items = []
        seen_ctrl_keys = set()
        seen_rule_keys = set()

        def add_ctrl(item_dict):
            key = f"{item_dict.get('location_tab')}::{item_dict.get('type')}::{item_dict.get('field_code')}"
            if key in seen_ctrl_keys:
                return
            seen_ctrl_keys.add(key)
            ctrl_items.append(item_dict)

        def add_rule(rule_dict):
            key = f"{rule_dict.get('location_tab')}::{rule_dict.get('rule_name')}"
            if key in seen_rule_keys:
                return
            seen_rule_keys.add(key)
            rule_items.append(rule_dict)

        # 1. Parse layout tabs recursively for Controls (Reports, RelationshipItems, AddIns)
        raw_tabs = ws.get("tabs", [])
        all_tabs = _collect_all_tabs(raw_tabs)
        tab_names = set()

        if all_tabs:
            for tab in all_tabs:
                t_label = (
                    tab.get("text") or
                    tab.get("text_label_name") or
                    tab.get("name") or
                    tab.get("label") or
                    "Form Tab"
                )

                # Analytics Reports (standalone <Report> controls e.g. AcId=9029)
                for rpt in tab.get("reports", []):
                    rid = str(rpt.get("id") or rpt.get("ac_id") or rpt.get("report_id") or rpt.get("search_report_id") or "")
                    if rid and rid != "None":
                        tab_names.add(t_label)
                        add_ctrl({
                            "type": "Report",
                            "location_tab": t_label,
                            "field_code": f"Report_{rid}",
                            "details": f"ReportId: {rid} | ExecuteOnNew: {rpt.get('execute_on_new', False)} | DelayExecution: {rpt.get('delay_report_execution', '')}"
                        })

                # BUI Add-Ins & Extensions (<AddInItem>)
                for ai in tab.get("add_ins", []) + tab.get("add_in_items", []):
                    ainame = ai.get("name") or ai.get("item_type") or ai.get("add_in_name") or "AddIn"
                    tab_names.add(t_label)
                    ai_tech = "BUI Extension" if ai.get("bui_extension") else "Desktop Add-In"
                    add_ctrl({
                        "type": "AddInItem",
                        "location_tab": t_label,
                        "field_code": ainame,
                        "details": f"Plugin: {ainame} | Type: {ai_tech} | FileId: {ai.get('file_id', 'N/A')} | Assembly: {ai.get('assembly', 'N/A')}"
                    })

                # Relationship Items & Sub-lists (<RelationshipItem>)
                for rel in tab.get("relationships", []) + tab.get("relationship_items", []):
                    r_type = rel.get("type") or rel.get("item_type") or "Relationship"
                    r_obj  = rel.get("object_id") or bound_obj
                    acid   = rel.get("ac_id")
                    srid   = rel.get("search_report_id")

                    tab_names.add(t_label)
                    details_parts = []
                    if r_obj and r_obj != bound_obj:
                        details_parts.append(f"Target Object: {r_obj}")
                    if acid and str(acid) != "None":
                        details_parts.append(f"Primary Report (AcId): {acid}")
                    if srid and str(srid) not in ("0", "None"):
                        details_parts.append(f"Search Report (SearchId): {srid}")

                    details_str = " | ".join(details_parts) if details_parts else "Relationship Sub-List Container"

                    add_ctrl({
                        "type": "RelationshipItem",
                        "location_tab": t_label,
                        "field_code": r_type,
                        "details": details_str
                    })

        # 2. Workspace Business Rules & Events (<Rule>)
        rules = ws.get("rules", [])
        for rule in rules:
            r_name = rule.get("name") or "Workspace Rule"
            trig_s, cond_s, act_s = _get_rule_components(rule)
            add_rule({
                "location_tab": "Workspace Rules",
                "rule_name": r_name,
                "trigger": trig_s,
                "condition": cond_s,
                "action": act_s
            })

        # 3. Object Business Rules matching bound_obj (<Rule>)
        for rule_set in rule_list:
            if str(rule_set.get("object", "")).lower() == bound_obj.lower():
                for r in rule_set.get("rules", []):
                    rname = r.get("name") or "Object Rule"
                    trig_s, cond_s, act_s = _get_rule_components(r)
                    add_rule({
                        "location_tab": "Object Rules",
                        "rule_name": rname,
                        "trigger": trig_s,
                        "condition": cond_s,
                        "action": act_s
                    })

        # Overview statistics
        rpt_count = len([i for i in ctrl_items if i["type"] == "Report"])
        rel_count = len([i for i in ctrl_items if i["type"] == "RelationshipItem"])
        bui_count = len([i for i in ctrl_items if i["type"] == "AddInItem"])
        rul_count = len(rule_items)

        # Calculate Internal vs External Custom Scripts for bound entity
        cs_int_count = 0
        cs_ext_count = 0
        for s in script_list:
            s_name = s.get("file_name") or s.get("name") or ""
            target_objs = [str(o).lower() for o in (s.get("target_objects") or [])]
            if (bound_obj.lower() in target_objs) or (bound_obj.lower() in s_name.lower()):
                if _is_external_script(s):
                    cs_ext_count += 1
                else:
                    cs_int_count += 1

        total_integrations = rpt_count + rel_count + bui_count + rul_count + cs_int_count + cs_ext_count

        overview_rows.append([
            ws_name, bound_obj, len(tab_names), total_integrations,
            rpt_count, rel_count, bui_count, rul_count, cs_int_count, cs_ext_count
        ])

        # Create Workspace Sheet
        sheet = wb.create_sheet(title=_safe_tab_name(ws_name))
        curr_row = 1

        # ── TABLE 1: Layout Controls / Components ─────────────────────────
        raw_ctrl_headers = [
            "Tab Name", "Control Type",
            "Control", "Control Details"
        ]
        ctrl_table_rows = []
        for item in ctrl_items:
            ctrl_table_rows.append([
                item["location_tab"],
                item["type"],
                item["field_code"],
                item["details"]
            ])

        active_ctrl_headers, active_ctrl_rows = prune_empty_table(raw_ctrl_headers, ctrl_table_rows)

        if active_ctrl_headers:
            for col_idx, h in enumerate(active_ctrl_headers, start=1):
                cell = sheet.cell(row=curr_row, column=col_idx, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = THIN_BORDER
            sheet.freeze_panes = "A2"
            sheet.row_dimensions[curr_row].height = 26
            curr_row += 1

            for r_idx, row_vals in enumerate(active_ctrl_rows, start=curr_row):
                fill = ALT_ROW_FILL if (r_idx % 2 == 0) else None
                for c_idx, val in enumerate(row_vals, start=1):
                    cell = sheet.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = BODY_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(vertical="center")
                    if fill:
                        cell.fill = fill
            curr_row += len(active_ctrl_rows)
            curr_row += 1  # Blank Row Separator between tables

        # ── TABLE 2: Workspace Rules (Only generated if rules exist) ──────
        if rule_items:
            # 1. Section Title Banner: "Workspace Rules"
            sheet.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=4)
            title_cell = sheet.cell(row=curr_row, column=1, value="Workspace Rules")
            title_cell.font = SECTION_FONT
            title_cell.fill = SECTION_FILL
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            sheet.row_dimensions[curr_row].height = 26
            curr_row += 1

            # 2. Table 2 Headers (4 Columns: Rule Name, Trigger, Condition, Action)
            rule_headers = ["Rule Name", "Trigger", "Condition", "Action"]
            for col_idx, h in enumerate(rule_headers, start=1):
                cell = sheet.cell(row=curr_row, column=col_idx, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = THIN_BORDER
            sheet.row_dimensions[curr_row].height = 26
            curr_row += 1

            # 3. Table 2 Data Rows
            for r_idx, r_item in enumerate(rule_items, start=curr_row):
                fill = ALT_ROW_FILL if (r_idx % 2 == 0) else None
                row_vals = [
                    r_item["rule_name"],
                    r_item["trigger"],
                    r_item["condition"],
                    r_item["action"]
                ]
                for c_idx, val in enumerate(row_vals, start=1):
                    cell = sheet.cell(row=r_idx, column=c_idx, value=val)
                    cell.font = BODY_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(vertical="center")
                    if fill:
                        cell.fill = fill
            curr_row += len(rule_items)

        # Set Column Dimensions
        sheet.column_dimensions["A"].width = 30  # Tab Name / Rule Name
        sheet.column_dimensions["B"].width = 28  # Control Type / Trigger
        sheet.column_dimensions["C"].width = 35  # Field ID / Condition
        sheet.column_dimensions["D"].width = 75  # Integration Details & Config / Action

    # Populate Overview Sheet with Column Pruning (if enabled)
    if include_overview:
        active_overview_headers, active_overview_rows = prune_empty_table(raw_overview_headers, overview_rows)

        overview_sheet = wb.create_sheet(title="Workspace_Overview")
        for col_idx, h in enumerate(active_overview_headers, start=1):
            cell = overview_sheet.cell(row=1, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        overview_sheet.freeze_panes = "A2"

        for r_idx, o_row in enumerate(active_overview_rows, start=2):
            fill = ALT_ROW_FILL if r_idx % 2 == 0 else None
            for c_idx, val in enumerate(o_row, start=1):
                cell = overview_sheet.cell(row=r_idx, column=c_idx, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                if fill:
                    cell.fill = fill

        for col_idx, h in enumerate(active_overview_headers, start=1):
            col_letter = get_column_letter(col_idx)
            overview_sheet.column_dimensions[col_letter].width = min(max(len(h) + 4, 18), 40)

    for s in wb.worksheets:
        if s.max_column > 0 and s.max_row > 0:
            max_c_letter = get_column_letter(s.max_column)
            s.print_area = f"A1:{max_c_letter}{s.max_row}"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path

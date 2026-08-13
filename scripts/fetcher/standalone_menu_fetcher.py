#!/usr/bin/env python3
"""
Standalone OSVC System Menu Fields Fetcher & Excel Generator.

Fetches OSVC standard system menu fields and namedID menu options highlighted in System Menus:
- Answer Access Levels
- Answer Statuses
- Channel Types
- Chat Agent Statuses
- Chat Queues
- Contact Roles
- Contact Types
- Incident Queues
- Incident Severities
- Incident Statuses
- Organization Address Types

Queries the OSVC Connect REST API via HTTP GET requests and generates a multi-tab Excel workbook.

Usage:
    python scripts/fetcher/standalone_menu_fetcher.py
"""

import os
import sys
import argparse
import requests
from requests.auth import HTTPBasicAuth
from urllib.parse import urlparse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure module import path includes current and parent directories
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
ROOT_DIR = os.path.dirname(PARENT_DIR)

for d in (SCRIPT_DIR, PARENT_DIR, ROOT_DIR):
    if d and d not in sys.path:
        sys.path.insert(0, d)

# Default connection settings
DEFAULT_HOST = os.environ.get("OSVC_HOST", "")
DEFAULT_USER = os.environ.get("OSVC_USERNAME", "")
DEFAULT_PASS = os.environ.get("OSVC_PASSWORD", "")

if not DEFAULT_HOST or not DEFAULT_USER or not DEFAULT_PASS:
    try:
        config_path = os.path.join(ROOT_DIR, "field_extractor", "config.py")
        if os.path.exists(config_path):
            import importlib.util
            spec = importlib.util.spec_from_file_location("field_extractor_config", config_path)
            cfg_mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(cfg_mod)
            DEFAULT_HOST = getattr(cfg_mod, "BASE_URL", DEFAULT_HOST)
            DEFAULT_USER = getattr(cfg_mod, "USERNAME", DEFAULT_USER)
            DEFAULT_PASS = getattr(cfg_mod, "PASSWORD", DEFAULT_PASS)
    except Exception:
        pass

DEFAULT_OUTPUT_PATH = os.path.join(SCRIPT_DIR, "results", "Standard_Menu_Fields.xlsx")

# ---------------------------------------------------------------------------
# Highlighted System Menus & Candidate Endpoint Registry
# ---------------------------------------------------------------------------

HIGHLIGHTED_SYSTEM_MENUS = [
    {
        "name": "Answer Access Levels",
        "endpoints": ["namedIDs/answers/accessLevels", "namedIDs/accessLevels", "accessLevels"]
    },
    {
        "name": "Answer Statuses",
        "endpoints": ["namedIDs/answers/statusWithType", "namedIDs/answers/status", "answerStatuses", "namedIDs/answerStatuses"]
    },
    {
        "name": "Channel Types",
        "endpoints": ["namedIDs/incidents/channel", "channelTypes", "namedIDs/channelTypes", "channels"]
    },
    {
        "name": "Chat Agent Statuses",
        "endpoints": ["namedIDs/chats/agentStatus", "chatAgentStatuses", "namedIDs/chatAgentStatuses"]
    },
    {
        "name": "Chat Queues",
        "endpoints": ["namedIDs/incidents/chatQueue", "chatQueues", "namedIDs/chatQueues"]
    },
    {
        "name": "Contact Roles",
        "endpoints": ["namedIDs/contacts/role", "contactRoles", "namedIDs/contactRoles"]
    },
    {
        "name": "Contact Types",
        "endpoints": ["namedIDs/contacts/contactType", "contactTypes", "namedIDs/contactTypes"]
    },
    {
        "name": "Incident Queues",
        "endpoints": ["namedIDs/incidents/queue", "queues", "namedIDs/queues"]
    },
    {
        "name": "Incident Severities",
        "endpoints": ["namedIDs/incidents/severity", "severities", "namedIDs/severities"]
    },
    {
        "name": "Incident Statuses",
        "endpoints": ["namedIDs/incidents/statusWithType", "namedIDs/incidents/status", "incidentStatuses", "namedIDs/incidentStatuses"]
    },
    {
        "name": "Organization Address Types",
        "endpoints": ["namedIDs/organizations/addresses/addressType", "organizationAddressTypes", "namedIDs/organizationAddressTypes"]
    }
]

# ---------------------------------------------------------------------------
# Excel Styling Definitions
# ---------------------------------------------------------------------------

HEADER_FILL   = PatternFill("solid", fgColor="2E75B6")   # Professional blue
SUMMARY_FILL  = PatternFill("solid", fgColor="1F4E78")   # Dark blue header
ALT_ROW_FILL  = PatternFill("solid", fgColor="F2F7FA")   # Soft light blue zebra
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT    = Font(name="Calibri", bold=True, color="1F4E78", size=14)
SUBTITLE_FONT = Font(name="Calibri", italic=True, color="595959", size=10)
BODY_FONT     = Font(name="Calibri", size=10)
BOLD_FONT     = Font(name="Calibri", bold=True, size=10)

THIN_BORDER   = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9")
)


def _clean_host_url(host):
    """Extracts scheme and netloc domain from any host string or URL."""
    host = host.strip()
    if not host.startswith('http://') and not host.startswith('https://'):
        host = f"https://{host}"
    parsed = urlparse(host)
    scheme = parsed.scheme or 'https'
    netloc = parsed.netloc or parsed.path.split('/')[0]
    return f"{scheme}://{netloc}"


def _safe_sheet_name(name):
    """Sanitizes string for a valid Excel sheet tab name (max 31 chars)."""
    clean_name = "".join([c if c.isalnum() or c in (" ", "_") else "_" for c in name])
    return clean_name[:31]


def _style_header_row(ws, headers, fill=HEADER_FILL):
    """Styles the header row of a worksheet with frozen panes and auto-filters."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"
    ws.row_dimensions[1].height = 28


def _auto_adjust_column_widths(ws, max_cols=None):
    """Adjusts column widths based on maximum string length per column."""
    num_cols = max_cols or ws.max_column
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)


def fetch_system_menu_fields(host, username, password, custom_endpoints=None, log_cb=print):
    """
    Connects to OSVC Connect REST API via HTTP GET requests and fetches items
    for highlighted System Menus and namedIDs endpoints.
    """
    base_url = _clean_host_url(host)
    api_root = f"{base_url}/services/rest/connect/v1.4"
    auth = HTTPBasicAuth(username, password)
    headers = {
        'Accept': 'application/json',
        'OSvC-CREST-Application-Context': 'System-Menu-Fetch'
    }

    session = requests.Session()

    menu_results = []
    log_cb(f"[START] Fetching System Menu fields from: {api_root}")

    targets = list(HIGHLIGHTED_SYSTEM_MENUS)
    if custom_endpoints:
        for ep in custom_endpoints:
            targets.append({
                "name": ep.split('/')[-1].capitalize(),
                "endpoints": [ep]
            })

    for target in targets:
        menu_name = target["name"]
        candidate_paths = target["endpoints"]

        log_cb(f"[INFO] Processing System Menu: {menu_name}")

        fetched = False
        for path in candidate_paths:
            url = f"{api_root}/{path}" if not path.startswith("http") else path
            try:
                log_cb(f"[STRICT GET ONLY] Requesting endpoint: {url}")
                resp = session.get(url, auth=auth, headers=headers, timeout=30)
                if resp.status_code == 200:
                    data = resp.json()
                    raw_items = data.get("items", []) or data.get("objects", [])
                    if not raw_items and isinstance(data, list):
                        raw_items = data

                    parsed_items = []
                    for item in raw_items:
                        if isinstance(item, dict):
                            item_id = item.get("id", "-")
                            lookup_name = item.get("lookupName") or item.get("name") or item.get("label") or str(item_id)
                            canonical_url = ""
                            for link in item.get("links", []):
                                if link.get("rel") == "canonical":
                                    canonical_url = link.get("href", "")
                                    break
                            parsed_items.append({
                                "id": item_id,
                                "lookup_name": lookup_name,
                                "canonical_url": canonical_url,
                                "raw": item
                            })
                        else:
                            parsed_items.append({
                                "id": item,
                                "lookup_name": str(item),
                                "canonical_url": "",
                                "raw": item
                            })

                    menu_results.append({
                        "menu_name": menu_name,
                        "endpoint_url": url,
                        "endpoint_path": path,
                        "status_code": 200,
                        "items_count": len(parsed_items),
                        "items": parsed_items
                    })

                    log_cb(f"[SUCCESS] Extracted {len(parsed_items)} option items for '{menu_name}' via {path}")
                    fetched = True
                    break

            except Exception as err:
                log_cb(f"[WARNING] Request failed for {url}: {err}")

        if not fetched:
            log_cb(f"[WARNING] Unable to fetch items for System Menu '{menu_name}' from any candidate endpoint.")
            menu_results.append({
                "menu_name": menu_name,
                "endpoint_url": f"{api_root}/{candidate_paths[0]}",
                "endpoint_path": candidate_paths[0],
                "status_code": 404,
                "items_count": 0,
                "items": []
            })

    return menu_results


def export_menus_to_excel(menu_results, output_filepath, host_url=""):
    """
    Exports the fetched System Menu fields into a multi-tab Excel workbook (.xlsx).
    Tabs generated:
    1. Summary - Metadata metrics & menu summary
    2. System_Menus_Overview - Master list of highlighted menus & item counts
    3. All_Menu_Options - Consolidated catalog of all option items across all menus
    4...N - Individual sheets for each highlighted System Menu
    """
    os.makedirs(os.path.dirname(os.path.abspath(output_filepath)), exist_ok=True)
    wb = openpyxl.Workbook()

    # ---------------------------------------------------------------------------
    # Sheet 1: Summary Page
    # ---------------------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    ws_summary.cell(row=1, column=1, value="OSVC System Menu Fields Extraction Summary").font = TITLE_FONT
    ws_summary.cell(row=2, column=1, value=f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = SUBTITLE_FONT

    tot_menus = len(menu_results)
    tot_items = sum(m["items_count"] for m in menu_results)
    successful_menus = len([m for m in menu_results if m["items_count"] > 0])

    metadata = [
        ("Host Endpoint URL", host_url or "N/A"),
        ("Highlighted System Menus", tot_menus),
        ("Menus with Configured Items", successful_menus),
        ("Total Menu Option Items Extracted", tot_items)
    ]

    ws_summary.cell(row=4, column=1, value="Metric").font = HEADER_FONT
    ws_summary.cell(row=4, column=1).fill = SUMMARY_FILL
    ws_summary.cell(row=4, column=2, value="Value").font = HEADER_FONT
    ws_summary.cell(row=4, column=2).fill = SUMMARY_FILL

    for idx, (m_name, m_val) in enumerate(metadata, start=5):
        cell_a = ws_summary.cell(row=idx, column=1, value=m_name)
        cell_b = ws_summary.cell(row=idx, column=2, value=m_val)
        cell_a.font = BOLD_FONT
        cell_b.font = BODY_FONT
        cell_a.border = THIN_BORDER
        cell_b.border = THIN_BORDER

    _auto_adjust_column_widths(ws_summary, max_cols=2)

    # ---------------------------------------------------------------------------
    # Sheet 2: System_Menus_Overview
    # ---------------------------------------------------------------------------
    ws_overview = wb.create_sheet(title="System_Menus_Overview")
    ws_overview.views.sheetView[0].showGridLines = True

    overview_headers = ["System Menu Name", "REST Endpoint Path", "Status Code", "Items Count", "Sample Option Values"]
    _style_header_row(ws_overview, overview_headers)

    for row_idx, menu_info in enumerate(menu_results, start=2):
        items = menu_info["items"]
        sample_str = ", ".join([it["lookup_name"] for it in items[:5]]) if items else "No items configured"

        row_data = [
            menu_info["menu_name"],
            menu_info["endpoint_path"],
            menu_info["status_code"],
            menu_info["items_count"],
            sample_str
        ]

        fill = ALT_ROW_FILL if row_idx % 2 == 0 else None
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws_overview.cell(row=row_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            if c_idx in (3, 4):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    _auto_adjust_column_widths(ws_overview)

    # ---------------------------------------------------------------------------
    # Sheet 3: All_Menu_Options Consolidated Catalog
    # ---------------------------------------------------------------------------
    ws_all_opts = wb.create_sheet(title="All_Menu_Options")
    ws_all_opts.views.sheetView[0].showGridLines = True

    all_opts_headers = ["System Menu Name", "Option ID", "Lookup Name", "Canonical Resource URL"]
    _style_header_row(ws_all_opts, all_opts_headers)

    all_opt_row = 2
    for menu_info in menu_results:
        m_name = menu_info["menu_name"]
        for item in menu_info["items"]:
            row_data = [
                m_name,
                item["id"],
                item["lookup_name"],
                item["canonical_url"] or "-"
            ]

            fill = ALT_ROW_FILL if all_opt_row % 2 == 0 else None
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws_all_opts.cell(row=all_opt_row, column=c_idx, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center")
                if fill:
                    cell.fill = fill
                if c_idx == 2:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            all_opt_row += 1

    _auto_adjust_column_widths(ws_all_opts)

    # ---------------------------------------------------------------------------
    # Sheets 4...N: Per System Menu Worksheets
    # ---------------------------------------------------------------------------
    for menu_info in menu_results:
        m_name = menu_info["menu_name"]
        sheet_title = _safe_sheet_name(m_name)

        ws_menu = wb.create_sheet(title=sheet_title)
        ws_menu.views.sheetView[0].showGridLines = True

        per_menu_headers = ["Option ID", "Lookup Name", "Canonical Resource URL"]
        _style_header_row(ws_menu, per_menu_headers)

        items = menu_info["items"]
        if not items:
            ws_menu.cell(row=2, column=1, value="No menu options configured or accessible at endpoint").font = Font(name="Calibri", italic=True, color="7F7F7F")
        else:
            for item_idx, item in enumerate(items, start=2):
                row_data = [
                    item["id"],
                    item["lookup_name"],
                    item["canonical_url"] or "-"
                ]

                fill = ALT_ROW_FILL if item_idx % 2 == 0 else None
                for c_idx, val in enumerate(row_data, start=1):
                    cell = ws_menu.cell(row=item_idx, column=c_idx, value=val)
                    cell.font = BODY_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(vertical="center")
                    if fill:
                        cell.fill = fill
                    if c_idx == 1:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

        _auto_adjust_column_widths(ws_menu)

    wb.save(output_filepath)
    print(f"[SUCCESS] System Menu Excel workbook successfully generated: {output_filepath}")


def main():
    parser = argparse.ArgumentParser(description="OSVC System Menu Fields Fetcher & Excel Generator")
    parser.add_argument("--host", default=DEFAULT_HOST, help="OSVC Host domain or full endpoint URL")
    parser.add_argument("--username", default=DEFAULT_USER, help="OSVC REST API Username")
    parser.add_argument("--password", default=DEFAULT_PASS, help="OSVC REST API Password")
    parser.add_argument("--output", default=DEFAULT_OUTPUT_PATH, help="Output Excel file path (.xlsx)")
    parser.add_argument("--endpoints", nargs="+", default=None, help="Additional custom REST endpoints/paths to fetch (e.g. namedIDs/incidents/assignedTo/staffGroup)")

    args = parser.parse_args()

    if not args.host or not args.username or not args.password:
        print("[ERROR] Missing OSVC connection credentials.")
        print("Please provide --host, --username, and --password as arguments or set environment variables OSVC_HOST, OSVC_USERNAME, OSVC_PASSWORD.")
        sys.exit(1)

    print("[START] Connecting to OSVC Connect REST API for System Menus...")
    print(f"[INFO] Target Host: {args.host}")
    print(f"[INFO] Username   : {args.username}")
    print(f"[INFO] Output Path: {args.output}")

    try:
        menu_results = fetch_system_menu_fields(
            host=args.host,
            username=args.username,
            password=args.password,
            custom_endpoints=args.endpoints,
            log_cb=print
        )

        print("[INFO] Generating System Menu Excel report...")
        export_menus_to_excel(menu_results, args.output, host_url=args.host)
        print(f"[DONE] Extraction completed successfully. Excel saved at: {args.output}")

    except Exception as err:
        print(f"[ERROR] System Menu extraction failed: {err}")
        sys.exit(1)


if __name__ == "__main__":
    main()

import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

HEADER_FILL  = PatternFill("solid", fgColor="2E75B6")   # professional blue
ALT_ROW_FILL = PatternFill("solid", fgColor="DEEAF1")   # light blue-grey
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Calibri", size=10)
THIN_BORDER  = Border(
    left=Side(style="thin", color="9DC3E6"),
    right=Side(style="thin", color="9DC3E6"),
    top=Side(style="thin", color="9DC3E6"),
    bottom=Side(style="thin", color="9DC3E6"),
)

def _style_sheet(ws, headers):
    """Apply header style, freeze the top row, and add auto-filter to a worksheet."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Keep header visible when scrolling
    ws.freeze_panes = "A2"

    # Add filter dropdowns on all header columns
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"

    ws.row_dimensions[1].height = 34

    # Auto-fit column widths (capped at 50)
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max(len(header) + 4, 14), 50)

def _write_rows(ws, rows, headers, start_row=2):
    """Write data rows with alternating row color."""
    for row_idx, row_data in enumerate(rows, start=start_row):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if fill:
                cell.fill = fill

def _safe_tab_name(name):
    """Sanitize a string to be a valid Excel sheet name (max 31 chars)."""
    name = re.sub(r'[\\/*?:\[\]]', '_', name)
    return name[:31]


# ---------------------------------------------------------------------------
# Option value formatting
# ---------------------------------------------------------------------------

_TRIGGER_MAP = {
    "onnew":    "New",
    "onedit":   "Edit",
    "onsave":   "Save",
    "onalways": "Always",
    "ondelete": "Delete",
}

def _format_option(raw):
    """
    Converts OSVC workspace option strings to clean readable values.
    Examples:
      ""                            -> No
      "OnNew:~any~;OnEdit:~any~"   -> Yes (New + Edit)
      "OnEdit:~any~"               -> Yes (Edit Only)
      "OnNew:~any~"                -> Yes (New Only)
      "True"                       -> Yes
      "False"                      -> No
    """
    if not raw or raw.strip() == "":
        return "No"

    raw = raw.strip()

    # Simple boolean-style values
    if raw.lower() in ("true", "yes", "1"):
        return "Yes"
    if raw.lower() in ("false", "no", "0"):
        return "No"

    # Parse segment list like OnNew:~any~;OnEdit:~any~
    segments = raw.split(";")
    triggers = []
    for seg in segments:
        seg = seg.strip()
        if not seg:
            continue
        if ":" in seg:
            trigger, condition = seg.split(":", 1)
            trigger   = trigger.strip().lower()
            condition = condition.strip().strip("~")
            label = _TRIGGER_MAP.get(trigger, trigger.capitalize())
            if condition == "any":
                triggers.append(label)
            else:
                triggers.append(f"{label} ({condition})")
        else:
            triggers.append(seg)

    if not triggers:
        return "No"
    if len(triggers) == 1:
        return f"Yes ({triggers[0]} Only)"
    return "Yes (" + " + ".join(triggers) + ")"


# ---------------------------------------------------------------------------
# Field key builder (PackageName$Name)
# ---------------------------------------------------------------------------

_SYSTEM_PACKAGES = {"oracleservicecloud", "rightnow", ""}

def _field_type_from_obj(field_dict):
    """
    Determines Field Type from the Object XML field definition.
    Primary signal: PackageName attribute — if it is a non-system package, the field is Custom.
    Secondary signal: whether the Name itself contains '$'.
    Falls back to IsSystemField attribute.

    Values returned:
      'Custom (PackageName)' — field created by the org (e.g. C, CO, MY_PKG)
      'System Field'         — built-in OSVC system field (IsSystemField=True)
      'OSVC Standard'        — built-in OSVC platform field (IsSystemField=False)
    """
    pkg  = (field_dict.get("package_name") or "").strip()
    name = (field_dict.get("field_name")   or "").strip()

    # Package name is the clearest signal
    if pkg and pkg.lower() not in _SYSTEM_PACKAGES:
        return f"Custom ({pkg})"

    # If the name itself contains $ it is a custom field (e.g. C$FieldName in raw XML)
    if "$" in name:
        pkg_part = name.split("$", 1)[0]
        return f"Custom ({pkg_part})"

    # Fall back to the IsSystemField flag from the XML
    return "System Field" if field_dict.get("is_system_field", False) else "OSVC Standard"


def _field_type_from_ws_id(raw_field_id):
    """
    Determines Field Type from the raw workspace FieldId string alone.
    If the FieldId contains '$' it is a custom field.
    Examples:
      'C$PhoneExt'                  -> Custom (C)
      'CustomFields.c$org_id_temp'  -> Custom (c)
      'Name.First'                  -> System Field
      'OrgId'                       -> System Field
    """
    if not raw_field_id:
        return "System Field"
    # Strip object and CustomFields prefix to get the bare field token
    token = re.sub(r'^[^.]+\.', '', raw_field_id)           # remove e.g. "Contact."
    token = re.sub(r'^CustomFields\.', '', token, flags=re.IGNORECASE)
    if "$" in token:
        pkg_part = token.split("$", 1)[0]
        return f"Custom ({pkg_part})"
    return "System Field"

def _obj_field_key(field_dict):
    """
    Builds a lookup key for an object field as PackageName$Name.
    System fields (OracleServiceCloud package) use just the field Name.
    Custom package fields use PackageName$Name (e.g. C$org_id_temp).
    """
    pkg  = (field_dict.get("package_name") or "").strip()
    name = (field_dict.get("field_name")   or "").strip()
    if pkg.lower() in _SYSTEM_PACKAGES:
        return name.lower()
    return f"{pkg}${name}".lower()


def _ws_field_key(raw_field_id):
    """
    Normalizes a workspace FieldId into the same key format used by _obj_field_key.
    Examples:
      "Name.First"                     -> "name.first"
      "C$PhoneExt"                     -> "c$phoneext"
      "CustomFields.c$org_id_temp"     -> "c$org_id_temp"
      "OrgId"                          -> "orgid"
    """
    if not raw_field_id:
        return ""
    key = raw_field_id.strip()
    # Strip leading ObjectId prefix (e.g. "Contact." when already prefixed)
    key = re.sub(r'^[A-Za-z0-9_]+\.[A-Za-z0-9_]+\.', '', key)  # e.g. Contact.CustomFields.
    # Collapse CustomFields. prefix
    key = re.sub(r'^CustomFields\.', '', key, flags=re.IGNORECASE)
    return key.lower()


# ---------------------------------------------------------------------------
# Object index builder
# ---------------------------------------------------------------------------

def _normalize_obj_keys(oname):
    """Returns singular and plural variations of an object name to bridge REST API and Workspace XML naming."""
    oname = oname.lower().strip()
    keys = {oname}
    if oname.endswith('s') and not oname.endswith('ss'):
        keys.add(oname[:-1])
    else:
        keys.add(oname + 's')
    if oname.endswith('ies'):
        keys.add(oname[:-3] + 'y')
    elif oname.endswith('y'):
        keys.add(oname[:-1] + 'ies')
    return keys

_FIELD_ALIASES = {
    "phoffice": "phones.office",
    "phmobile": "phones.mobile",
    "phhome":   "phones.home",
    "phfax":    "phones.fax",
    "addr":     "address",
    "state":    "address.stateorprovince",
    "city":     "address.city",
    "postalcode": "address.postalcode",
}

def _build_obj_index(objects_map):
    """
    Returns dict keyed by lowercase object name (singular & plural) ->
    {field_key: field_dict} where field_key = PackageName$Name, plain Name, or token.
    """
    indexed = {}
    for oname, odata in (objects_map or {}).items():
        lookup = {}
        for of in odata.get("fields", []):
            key = _obj_field_key(of)
            if key:
                lookup[key] = of

            plain = (of.get("field_name") or "").lower()
            if plain:
                lookup[plain] = of

            field_id = (of.get("field_id") or "").lower()
            if field_id:
                lookup[field_id] = of

            # Fallback for composite names (e.g. name.first -> first)
            if "." in plain:
                last_part = plain.split(".")[-1]
                if last_part not in lookup:
                    lookup[last_part] = of

            # Fallback for custom fields (e.g. c$phone_ext -> phone_ext)
            if "$" in key:
                bare_custom = key.split("$")[-1]
                if bare_custom not in lookup:
                    lookup[bare_custom] = of

        # Add field alias entries (e.g. phoffice -> phones.office)
        for alias_key, target_field_name in _FIELD_ALIASES.items():
            if target_field_name in lookup and alias_key not in lookup:
                lookup[alias_key] = lookup[target_field_name]

        for k in _normalize_obj_keys(oname):
            if k not in indexed:
                indexed[k] = lookup
            else:
                indexed[k].update(lookup)

    return indexed


# ---------------------------------------------------------------------------
# Workspace field enrichment
# ---------------------------------------------------------------------------

def _enrich_workspace_fields(ws_fields, objects_map, bound_object, standard_objects_map=None, custom_objects_map=None):
    """
    Enriches workspace fields with object schema metadata (Max Length, Data Type, Is Nullable, Is Lookup, etc.).
    First differentiates workspace fields into Standard vs Custom:
      - Standard fields -> searched in standard_objects_map (REST API schemas)
      - Custom fields / Custom objects -> searched in custom_objects_map (XML schemas)
    """
    if standard_objects_map is None and custom_objects_map is None:
        std_map = {}
        cst_map = {}
        for oname, odata in (objects_map or {}).items():
            if '.' in oname:
                cst_map[oname] = odata
            else:
                std_map[oname] = odata
                # Standard objects may also contain custom fields
                c_fields = [f for f in odata.get("fields", []) if f.get("package_name", "").lower() not in _SYSTEM_PACKAGES or "$" in f.get("field_name", "")]
                if c_fields:
                    cst_map[oname] = {"object_name": odata.get("object_name", oname), "fields": c_fields}
        std_indexed = _build_obj_index(std_map)
        cst_indexed = _build_obj_index(cst_map)
    else:
        std_indexed = _build_obj_index(standard_objects_map or {})
        cst_indexed = _build_obj_index(custom_objects_map or {})

    combined_indexed = _build_obj_index(objects_map or {})
    enriched = []

    for wf in ws_fields:
        target_obj = wf.get("target_object") or bound_object
        target_key = target_obj.lower()

        raw_id     = wf.get("raw_field_id", "") or wf.get("field_code", "")
        f_code     = wf.get("field_code", "")

        ws_key     = _ws_field_key(raw_id)
        ws_key_alt = _ws_field_key(f_code)

        # Differentiate Standard vs Custom field
        is_custom_ws = ("$" in raw_id) or ("customfields" in raw_id.lower()) or ("$" in f_code)

        # Select primary index based on standard vs custom differentiation
        if is_custom_ws:
            primary_idx   = cst_indexed.get(target_key, {})
            secondary_idx = std_indexed.get(target_key, {})
        else:
            primary_idx   = std_indexed.get(target_key, {})
            secondary_idx = cst_indexed.get(target_key, {})

        comb_idx = combined_indexed.get(target_key, {})

        matched = (
            primary_idx.get(ws_key) or primary_idx.get(ws_key_alt) or
            secondary_idx.get(ws_key) or secondary_idx.get(ws_key_alt) or
            comb_idx.get(ws_key) or comb_idx.get(ws_key_alt)
        )

        # Fallback to trailing component match (e.g. name.first -> first)
        if not matched and "." in ws_key:
            bare = ws_key.split(".")[-1]
            matched = primary_idx.get(bare) or secondary_idx.get(bare) or comb_idx.get(bare)

        if matched:
            data_type     = matched.get("data_type", "Text")
            ftype         = _field_type_from_obj(matched)
            is_nullable   = "Yes" if matched.get("is_nullable") else "No"
            is_lookup     = "Yes" if matched.get("is_lookup")   else "No"
            is_list       = "Yes" if matched.get("is_list")     else "No"
            is_autoupdate = "Yes" if matched.get("is_autoupdate") else "No"
            is_sequence   = "Yes" if matched.get("is_sequence")   else "No"
            max_len       = str(matched.get("max_length", "-"))
            desc          = matched.get("description", "")
            avail_get     = "Yes" if matched.get("is_available_get", True) else "No"
            avail_post    = "Yes" if matched.get("is_available_post", False) else "No"
            avail_patch   = "Yes" if matched.get("is_available_patch", False) else "No"
            is_deprec     = "Yes" if matched.get("is_deprecated", False) else "No"

            pkg   = (matched.get("package_name") or "").strip()
            fname = matched.get("field_name", "")
            if pkg.lower() in _SYSTEM_PACKAGES:
                obj_field_key = fname
            else:
                obj_field_key = f"{pkg}${fname}"
        else:
            data_type     = "Standard Data Field"
            ftype         = _field_type_from_ws_id(raw_id or f_code)
            is_nullable   = "Yes"
            is_lookup     = "Yes" if ("Name" in f_code or "Id" in f_code) else "No"
            is_list       = "No"
            is_autoupdate = "No"
            is_sequence   = "No"
            max_len       = "-"
            desc          = ""
            avail_get     = "Yes"
            avail_post    = "No"
            avail_patch   = "No"
            is_deprec     = "No"
            obj_field_key = raw_id

        item = dict(wf)
        item.update({
            "target_object":  target_obj,
            "obj_field_key":  obj_field_key,
            "data_type":      data_type,
            "field_type":     ftype,
            "is_nullable":    is_nullable,
            "is_lookup":      is_lookup,
            "is_list":        is_list,
            "is_autoupdate":  is_autoupdate,
            "is_sequence":    is_sequence,
            "max_length":     max_len,
            "description":    desc,
            "avail_get":      avail_get,
            "avail_post":     avail_post,
            "avail_patch":    avail_patch,
            "is_deprecated":  is_deprec,
            "required_fmt":   _format_option(wf.get("required_option", "")),
            "readonly_fmt":   _format_option(wf.get("readonly_option", "")),
        })
        enriched.append(item)

    return enriched


# ---------------------------------------------------------------------------
# Public Excel writers
# ---------------------------------------------------------------------------

def write_workspaces_excel(parsed_workspaces, objects_map, output_path):
    """
    workspaces.xlsx — one tab per workspace.
    Tab name = workspace name. No 'Workspace Name' column.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    headers = [
        "Bound Object", "Target Object", "Object Field Name",
        "Field Label", "Location / Tab", "Required", "Read Only",
        "Data Type", "Field Type", "Is Nullable", "Is Lookup", "Is List", "Is Auto Update", "Max Length",
        "Is Available GET", "Is Available POST", "Is Available PATCH", "Is Deprecated", "Description"
    ]

    for ws_data in parsed_workspaces:
        ws_name      = ws_data["workspace_name"]
        bound_object = ws_data.get("bound_object", "Contact")
        enriched     = _enrich_workspace_fields(ws_data.get("fields", []), objects_map, bound_object)

        sheet = wb.create_sheet(title=_safe_tab_name(ws_name))
        _style_sheet(sheet, headers)

        rows = []
        for item in enriched:
            rows.append([
                item["bound_object"],
                item["target_object"],
                item["obj_field_key"],
                item["field_label"],
                item["location_tab"],
                item["required_fmt"],
                item["readonly_fmt"],
                item["data_type"],
                item["field_type"],
                item["is_nullable"],
                item["is_lookup"],
                item["is_list"],
                item["is_autoupdate"],
                item["max_length"],
                item.get("avail_get", "Yes"),
                item.get("avail_post", "No"),
                item.get("avail_patch", "No"),
                item.get("is_deprecated", "No"),
                item.get("description", ""),
            ])

        _write_rows(sheet, rows, headers)

    wb.save(output_path)
    return output_path


def write_objects_excel(objects_map, output_path):
    """
    objects.xlsx / standard_objects.xlsx / custom_objects.xlsx — one tab per object.
    Tab name = object name. Field key shown as PackageName$Name.
    Outputs STRICTLY the 21 requested columns in exact order.
    """
    if not objects_map:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "No_Objects"
        ws.cell(row=1, column=1, value="No object schemas extracted.")
        wb.save(output_path)
        return output_path

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    headers = [
        "Field Key", "Field Label", "Data Type", "Field Type",
        "Is System Field", "Package Name", "Is Nullable", "Is Lookup",
        "Is Read Only", "Max Length", "Description",
        "Is Available GET", "Is Available POST", "Is Available PATCH",
        "Is Deprecated", "Is Enumerable", "Minimum", "Maximum",
        "$Ref", "Items", "Pattern"
    ]

    for obj_name, obj_data in objects_map.items():
        display_name = obj_data.get("object_name", obj_name)
        sheet = wb.create_sheet(title=_safe_tab_name(display_name))
        _style_sheet(sheet, headers)

        rows = []
        for of in obj_data.get("fields", []):
            key = _obj_field_key(of)

            is_enum_val = of.get("isEnumerable")
            if isinstance(is_enum_val, bool):
                is_enum_str = "Yes" if is_enum_val else "No"
            else:
                is_enum_str = str(is_enum_val) if is_enum_val is not None else "-"

            items_val = of.get("items")
            if isinstance(items_val, (dict, list)):
                items_str = str(items_val)
            elif of.get("is_list"):
                items_str = "IsList: Yes"
            else:
                items_str = str(items_val) if items_val is not None else "-"

            ref_val = of.get("$ref") or of.get("ref") or of.get("ref_url") or "-"

            row = [
                key,
                of.get("field_label", ""),
                of.get("data_type", ""),
                _field_type_from_obj(of),
                "Yes" if of.get("is_system_field") else "No",
                of.get("package_name", ""),
                "Yes" if of.get("is_nullable") else "No",
                "Yes" if of.get("is_lookup")   else "No",
                "Yes" if of.get("is_readonly") else "No",
                str(of.get("max_length", "-")),
                of.get("description", ""),
                "Yes" if of.get("is_available_get", True) else "No",
                "Yes" if of.get("is_available_post", False) else "No",
                "Yes" if of.get("is_available_patch", False) else "No",
                "Yes" if of.get("is_deprecated", False) else "No",
                is_enum_str,
                str(of.get("minimum", "-")),
                str(of.get("maximum", "-")),
                str(ref_val),
                items_str,
                str(of.get("pattern", "-"))
            ]
            rows.append(row)

        _write_rows(sheet, rows, headers)

    wb.save(output_path)
    return output_path


def write_combined_excel(parsed_workspaces, objects_map, output_path):
    """
    combined.xlsx — one tab per workspace.
    Tab name = workspace name. Fields enriched with object schema data.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    headers = [
        "Bound Object", "Target Object", "Object Field Name",
        "Field Label", "Workspace Tab", "Required", "Read Only",
        "Data Type", "Field Type", "Is Nullable",
        "Is Lookup", "Is List", "Is Auto Update", "Max Length", "Is Available GET", "Is Available POST",
        "Is Available PATCH", "Is Deprecated", "Description", "In Workspace Layout"
    ]

    for ws_data in parsed_workspaces:
        ws_name      = ws_data["workspace_name"]
        bound_object = ws_data.get("bound_object", "Contact")
        enriched     = _enrich_workspace_fields(ws_data.get("fields", []), objects_map, bound_object)

        sheet = wb.create_sheet(title=_safe_tab_name(ws_name))
        _style_sheet(sheet, headers)

        rows = []
        for item in enriched:
            rows.append([
                item["bound_object"],
                item["target_object"],
                item["obj_field_key"],
                item["field_label"],
                item["location_tab"],
                item["required_fmt"],
                item["readonly_fmt"],
                item["data_type"],
                item["field_type"],
                item["is_nullable"],
                item["is_lookup"],
                item["is_list"],
                item["is_autoupdate"],
                item["max_length"],
                item.get("avail_get", "Yes"),
                item.get("avail_post", "No"),
                item.get("avail_patch", "No"),
                item.get("is_deprecated", "No"),
                item.get("description", ""),
                "Yes",
            ])

        _write_rows(sheet, rows, headers)

    wb.save(output_path)
    return output_path
